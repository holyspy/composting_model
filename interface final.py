import customtkinter as ctk
from tkinter import messagebox
import numpy as np
import matplotlib
# Configurer explicitement le backend pour matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import plotly.graph_objects as go
import plotly.io as pio
from plotly.subplots import make_subplots
import math
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import webbrowser
import os
import tempfile
import signal
import sys
# Import pour les filtres et analyses de signal
from scipy.signal import savgol_filter
# Import du module de modèle
from modelvic import SimulationModel
import matplotlib.patches as mpatches

# Gestionnaire de signal pour capturer les interruptions
def signal_handler(sig, frame):
    print('Programme interrompu par l\'utilisateur')
    plt.close('all')
    sys.exit(0)

# Enregistrer le gestionnaire de signal
signal.signal(signal.SIGINT, signal_handler)

class BioProcessApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Simulation de Procédé Biologique")
        self.geometry("1280x800")
        
        # Set appearance mode and default color theme
        ctk.set_appearance_mode("System")
        ctk.set_default_color_theme("blue")
        
        # Initialize application data
        self.NS = 0
        self.substrates = []
        self.current_substrate = 0
        self.input_data = {}
        self.results = {}
        
        # Create tabs
        self.notebook = ctk.CTkTabview(self)
        self.notebook.pack(expand=True, fill="both", padx=15, pady=15)
        
        self.substrate_frame = self.notebook.add("Substrats")
        self.data_frame = self.notebook.add("Données")
        self.simulation_frame = self.notebook.add("Simulation")
        self.graph_frame = self.notebook.add("Graphiques")
        self.sensitivity_frame = self.notebook.add("Analyse Sensibilité")
        
        self.create_substrate_tab()
        self.create_data_tab()
        self.create_simulation_tab()
        self.create_graph_tab()
        self.create_sensitivity_tab()
        
        # Configurer le gestionnaire de fermeture
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        
    def on_closing(self):
        """Gérer proprement la fermeture de l'application"""
        try:
            # Libérer les ressources matplotlib
            plt.close('all')
            
            # Détruire les canvas
            if hasattr(self, 'solid_canvas'):
                self.solid_canvas.get_tk_widget().destroy()
            if hasattr(self, 'liquid_canvas'):
                self.liquid_canvas.get_tk_widget().destroy()
            if hasattr(self, 'gas_canvas'):
                self.gas_canvas.get_tk_widget().destroy()
            
            # Détruire l'application
            self.quit()
            self.destroy()
            
            # Forcer la fermeture de Python
            import sys
            sys.exit(0)
        except Exception as e:
            print(f"Erreur lors de la fermeture: {str(e)}")
            self.destroy()
            import sys
            sys.exit(1)
    
    def create_substrate_tab(self):
        container = ctk.CTkFrame(self.substrate_frame)
        container.pack(pady=20)
        
        # Input for number of substrates
        ctk.CTkLabel(container, text="Nombre de substrats:").grid(row=0, column=0)
        self.ns_entry = ctk.CTkEntry(container)
        self.ns_entry.grid(row=0, column=1)
        ctk.CTkButton(container, text="Valider", command=self.init_substrates).grid(row=0, column=2)
        
        # Container for substrate inputs
        self.substrate_inputs = ctk.CTkFrame(container)
        self.substrate_inputs.grid(row=1, column=0, columnspan=3, pady=20)
        
    def init_substrates(self):
        try:
            self.NS = int(self.ns_entry.get())
            self.create_dynamic_inputs()
        except ValueError:
            messagebox.showerror("Erreur", "Veuillez entrer un nombre valide")
    
    def create_dynamic_inputs(self):
        # Remove old inputs
        for widget in self.substrate_inputs.winfo_children():
            widget.destroy()
            
        # Main grid container
        main_grid = ctk.CTkFrame(self.substrate_inputs)
        main_grid.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Title and information
        title_frame = ctk.CTkFrame(main_grid)
        title_frame.grid(row=0, column=0, columnspan=6, sticky="ew", pady=(0, 10))
        
        substrate_title = ctk.CTkLabel(
            title_frame, 
            text=f"Substrat {self.current_substrate+1}/{self.NS}",
            font=("Arial", 14, "bold")
        )
        substrate_title.pack(pady=5)
        
        # Column headers
        headers = ["Paramètre", "Valeur", "Paramètre", "Valeur"]
        for i, header in enumerate(headers):
            ctk.CTkLabel(main_grid, text=header, font=("Arial", 12, "bold")).grid(
                row=1, column=i, padx=5, pady=5, sticky="ew"
            )
        
        # Basic parameters (first column)
        basic_params = [
            "Nom du substrat:", 
            "Température (°C):",
            "Débit (kg/h):", 
            "Fraction solide (%):",
            "Fraction solide volatile (%):", 
            "Fraction biologique (%):",
        ]
        
        # Valeurs par défaut pour les deux substrats
        default_values = [
            # Substrat 1
            {
                "name": "Substrat 1",
                "T": 22,
                "FR": 10000,
                "FS": 40,
                "FVS": 90,
                "FBVS": 90,
                "FfBVS": 80,
                "fKT20": 0.01,
                "sKT20": 0.005,
                "Cp": 0.9,
                "comp": [18, 26, 10, 1]
            },
            # Substrat 2
            {
                "name": "Substrat 2",
                "T": 22,
                "FR": 5000,
                "FS": 80.62,
                "FVS": 30,
                "FBVS": 60,
                "FfBVS": 60,
                "fKT20": 0.01,
                "sKT20": 0.005,
                "Cp": 2.5,
                "comp": [27, 38, 16, 1]
            }
        ]
        
        # Sélectionner les valeurs par défaut pour le substrat actuel
        if self.current_substrate < len(default_values):
            default_substrate = default_values[self.current_substrate]
        else:
            default_substrate = default_values[0]  # Utiliser le premier substrat comme défaut pour les autres
        
        self.entries = []
        
        # First column of parameters
        for i, label in enumerate(basic_params):
            ctk.CTkLabel(main_grid, text=label).grid(
                row=i+2, column=0, sticky="w", padx=5, pady=2
            )
            entry = ctk.CTkEntry(main_grid, width=120)
            entry.grid(row=i+2, column=1, padx=5, pady=2, sticky="ew")
            
            # Pré-remplir avec la valeur par défaut
            if i == 0:  # Nom du substrat
                entry.insert(0, default_substrate["name"])
            elif i == 1:  # Température
                entry.insert(0, str(default_substrate["T"]))
            elif i == 2:  # Débit
                entry.insert(0, str(default_substrate["FR"]))
            elif i == 3:  # Fraction solide
                entry.insert(0, str(default_substrate["FS"]))
            elif i == 4:  # Fraction solide volatile
                entry.insert(0, str(default_substrate["FVS"]))
            elif i == 5:  # Fraction biologique
                entry.insert(0, str(default_substrate["FBVS"]))
                
            self.entries.append(entry)
        
        # Second column of parameters
        advanced_params = [
            "Fraction rapide (%):", 
            "fKT20:", 
            "sKT20:", 
            "Capacité calorifique (kJ/kg°C):",
        ]
        
        for i, label in enumerate(advanced_params):
            ctk.CTkLabel(main_grid, text=label).grid(
                row=i+2, column=2, sticky="w", padx=5, pady=2
            )
            entry = ctk.CTkEntry(main_grid, width=120)
            entry.grid(row=i+2, column=3, padx=5, pady=2, sticky="ew")
            
            # Pré-remplir avec la valeur par défaut
            if i == 0:  # Fraction rapide
                entry.insert(0, str(default_substrate["FfBVS"]))
            elif i == 1:  # fKT20
                entry.insert(0, str(default_substrate["fKT20"]))
            elif i == 2:  # sKT20
                entry.insert(0, str(default_substrate["sKT20"]))
            elif i == 3:  # Capacité calorifique
                entry.insert(0, str(default_substrate["Cp"]))
                
            self.entries.append(entry)
            
        # Chemical composition section
        comp_row = len(basic_params) + 2  # Start after basic parameters
        
        comp_title = ctk.CTkLabel(
            main_grid, 
            text="Composition chimique (C₁H₂O₃N₄):", 
            font=("Arial", 12, "bold")
        )
        comp_title.grid(row=comp_row, column=0, columnspan=4, sticky="w", padx=5, pady=(10, 5))
        
        comp_row += 1
        comp_frame = ctk.CTkFrame(main_grid)
        comp_frame.grid(row=comp_row, column=0, columnspan=4, sticky="ew", padx=5, pady=5)
        
        # Create boxes for a, b, c, d horizontally
        self.comp_entries = []
        comp_labels = ["a (C):", "b (H):", "c (O):", "d (N):"]
        
        for i, label in enumerate(comp_labels):
            ctk.CTkLabel(comp_frame, text=label).grid(row=0, column=i*2, padx=5, pady=5, sticky="e")
            entry = ctk.CTkEntry(comp_frame, width=70)
            entry.grid(row=0, column=i*2+1, padx=5, pady=5, sticky="w")
            
            # Pré-remplir avec la valeur par défaut
            entry.insert(0, str(default_substrate["comp"][i]))
            
            self.comp_entries.append(entry)
            
        # Configure columns to expand evenly
        for i in range(8):
            comp_frame.grid_columnconfigure(i, weight=1)
            
        # Navigation buttons
        nav_row = comp_row + 1
        nav_frame = ctk.CTkFrame(main_grid)
        nav_frame.grid(row=nav_row, column=0, columnspan=4, pady=15, sticky="ew")
        
        if self.current_substrate > 0:
            ctk.CTkButton(nav_frame, text="Précédent", command=self.prev_substrate).pack(side="left", padx=5)
            
        if self.current_substrate < self.NS - 1:
            ctk.CTkButton(nav_frame, text="Suivant", command=self.next_substrate).pack(side="left", padx=5)
        else:
            ctk.CTkButton(nav_frame, text="Terminer", command=self.save_substrates).pack(side="left", padx=5)
            
        # Configure main grid columns to expand
        for i in range(4):
            main_grid.grid_columnconfigure(i, weight=1)
    
    def next_substrate(self):
        self.save_current_substrate()
        self.current_substrate += 1
        self.create_dynamic_inputs()
    
    def prev_substrate(self):
        self.save_current_substrate()
        self.current_substrate -= 1
        self.create_dynamic_inputs()
    
    def save_current_substrate(self):
        try:
            data = {
                "name": self.entries[0].get(),
                "T": float(self.entries[1].get()),
                "FR": float(self.entries[2].get()),
                "FS": float(self.entries[3].get()),
                "FVS": float(self.entries[4].get()),
                "FBVS": float(self.entries[5].get()),
                "FfBVS": float(self.entries[6].get()),
                "fKT20": float(self.entries[7].get()),
                "sKT20": float(self.entries[8].get()),
                "Cp": float(self.entries[9].get()),
                "composition": [
                    float(self.comp_entries[0].get()),  # a - carbone
                    float(self.comp_entries[1].get()),  # b - hydrogène
                    float(self.comp_entries[2].get()),  # c - oxygène
                    float(self.comp_entries[3].get())   # d - azote
                ]
            }
            if len(self.substrates) <= self.current_substrate:
                self.substrates.append(data)
            else:
                self.substrates[self.current_substrate] = data
        except ValueError as e:
            messagebox.showerror("Erreur", f"Veuillez entrer des valeurs numériques valides pour tous les champs.\nErreur: {str(e)}")
    
    def save_substrates(self):
        self.save_current_substrate()
        self.show_data_summary()
        messagebox.showinfo("Succès", "Tous les substrats ont été enregistrés!")
        self.notebook.set("Données")
    
    def create_data_tab(self):
        """Create a tab for displaying data in tabular format"""
        container = ctk.CTkFrame(self.data_frame)
        container.pack(expand=True, fill="both", padx=20, pady=20)
        
        # Create a frame for controls at the top
        controls_frame = ctk.CTkFrame(container)
        controls_frame.pack(fill="x", padx=10, pady=10)
        
        # Add an export button
        export_btn = ctk.CTkButton(
            controls_frame, 
            text="Exporter les données", 
            command=self.export_data
        )
        export_btn.pack(side="left", padx=10)
        
        # Add a refresh button
        refresh_btn = ctk.CTkButton(
            controls_frame, 
            text="Rafraîchir", 
            command=self.update_data_display
        )
        refresh_btn.pack(side="left", padx=10)
        
        # Create text area with scrollbar for data display
        text_frame = ctk.CTkFrame(container)
        text_frame.pack(expand=True, fill="both", padx=10, pady=10)
        
        # Add scrollbars
        y_scrollbar = ctk.CTkScrollbar(text_frame)
        y_scrollbar.pack(side="right", fill="y")
        
        x_scrollbar = ctk.CTkScrollbar(text_frame, orientation="horizontal")
        x_scrollbar.pack(side="bottom", fill="x")
        
        # Create the textbox with both scrollbars
        self.data_text = ctk.CTkTextbox(
            text_frame, 
            wrap="none",
            yscrollcommand=y_scrollbar.set,
            xscrollcommand=x_scrollbar.set,
            font=("Courier New", 12)
        )
        self.data_text.pack(expand=True, fill="both", padx=5, pady=5)
        
        # Configure scrollbars
        y_scrollbar.configure(command=self.data_text.yview)
        x_scrollbar.configure(command=self.data_text.xview)
    
    def export_data(self):
        """Export displayed data to an Excel file"""
        import tkinter.filedialog as filedialog
        import pandas as pd
        import os
        
        # Vérifier si pandas et openpyxl sont installés
        try:
            import pandas as pd
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            if messagebox.askyesno("Modules manquants", 
                               "Cette fonctionnalité nécessite les modules pandas et openpyxl. "
                               "Voulez-vous les installer automatiquement?"):
                try:
                    import subprocess
                    subprocess.check_call(["pip", "install", "pandas", "openpyxl"])
                    import pandas as pd
                    import openpyxl
                    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                    from openpyxl.utils import get_column_letter
                    messagebox.showinfo("Installation réussie", 
                                       "Modules installés avec succès!")
                except Exception as install_error:
                    messagebox.showerror("Erreur d'installation", 
                                        f"Impossible d'installer les modules: {str(install_error)}")
                    return
            else:
                return
        
        # Vérifier si des simulations existent
        if not hasattr(self, 'simulations') or not self.simulations:
            messagebox.showinfo("Information", "Aucune simulation à exporter. Lancez d'abord une simulation.")
            return
        
        # Ask for file location
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx"), ("Tous les fichiers", "*.*")],
            title="Exporter les données vers Excel"
        )
        
        if not file_path:
            return  # User cancelled
        
        try:
            # Créer un writer Excel avec openpyxl comme moteur
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # 1. Feuille des paramètres de simulation
                params_data = []
                for i, sim in enumerate(self.simulations):
                    params_data.append([
                        f"Simulation {i+1}: {sim['name']}",
                        sim['params']['HRT'],                   # HRT
                        sim['params']['air_flow'],              # Débit d'air
                        sim['params']['relative_humidity'],     # Humidité relative
                        sim['params']['ambient_temp'],          # Température ambiante
                        sim['params']['water_flow'],            # Débit d'eau
                        sim['params']['water_temp'],            # Température eau
                        "Oui" if sim['params']['air_alternance'] else "Non", # Alternance
                        sim['params']['air_on_time'],           # Temps ON
                        sim['params']['air_off_time']           # Temps OFF
                    ])
                
                # Créer le DataFrame avec les colonnes appropriées
                columns = [
                    "Simulation", "HRT (heures)", "Débit d'air (m³/h)", 
                    "Humidité relative (%)", "Température ambiante (°C)",
                    "Débit d'eau (kg/h)", "Température eau (°C)", 
                    "Alternance activée", "Temps ON (h)", "Temps OFF (h)"
                ]
                params_df = pd.DataFrame(params_data, columns=columns)
                params_df.to_excel(writer, sheet_name="Paramètres", index=False)
                
                # 2. Feuille des résultats finaux
                results_data = []
                for i, sim in enumerate(self.simulations):
                    data = sim["data"]
                    results_data.append([
                        f"Simulation {i+1}: {sim['name']}",
                        data["Temperatures"][0],                # Température initiale
                        data["Temperatures"][-1],               # Température finale
                        data["MoistureFraction"][0],            # Humidité initiale
                        data["MoistureFraction"][-1],           # Humidité finale
                        data["Solids"][0],                      # Solides initiaux
                        data["Solids"][-1],                     # Solides finaux
                        data.get("process_volume", "N/A")       # Volume du processus
                    ])
                
                # Créer le DataFrame avec les colonnes appropriées
                columns = [
                    "Simulation", "Température initiale (°C)", "Température finale (°C)", 
                    "Humidité initiale (%)", "Humidité finale (%)",
                    "Solides initiaux (kg)", "Solides finaux (kg)", 
                    "Volume du processus (m³)"
                ]
                results_df = pd.DataFrame(results_data, columns=columns)
                results_df.to_excel(writer, sheet_name="Résultats finaux", index=False)
                
                # 3. Feuille pour chaque simulation avec évolution temporelle
                for i, sim in enumerate(self.simulations):
                    data = sim["data"]
                    name = sim["name"]
                    sheet_name = f"Sim{i+1}_{name}"[:31]  # Excel limite les noms à 31 caractères
                    
                    # Créer un DataFrame pour toutes les données temporelles
                    time_data = {
                        "Temps (h)": data["Times"],
                        "Température (°C)": data["Temperatures"],
                        "Humidité (%)": data["MoistureFraction"],
                        "Solides (kg)": data["Solids"],
                        "Débit gaz (kg/h)": data["QExhaustgases"],
                        "Volume gaz (m³/h)": data["VExhaustgases"],
                        "Humidité relative (%)": data["RelativeHumidity"]
                    }
                    
                    time_df = pd.DataFrame(time_data)
                    time_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Récupérer le classeur pour appliquer le style
                workbook = writer.book
                
                # Définir les styles
                title_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
                header_font = Font(name='Arial', size=11, bold=True)
                
                title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_fill = PatternFill(start_color='95B3D7', end_color='95B3D7', fill_type='solid')
                alt_row_fill = PatternFill(start_color='E6EFF7', end_color='E6EFF7', fill_type='solid')
                
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Appliquer les styles aux feuilles
                for sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    
                    # Ajouter un titre et fusionner les cellules
                    ws.insert_rows(1)
                    title_text = f"Données de {sheet_name}"
                    ws.cell(row=1, column=1, value=title_text)
                    last_col = len(ws[2])  # Nombre de colonnes dans la ligne 2 (en-têtes)
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
                    
                    # Appliquer le style au titre
                    ws.cell(row=1, column=1).font = title_font
                    ws.cell(row=1, column=1).fill = title_fill
                    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Ajuster la hauteur de la première ligne
                    ws.row_dimensions[1].height = 25
                    
                    # Appliquer le style aux en-têtes (maintenant ligne 3)
                    for col in range(1, last_col + 1):
                        cell = ws.cell(row=2, column=col)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_align
                        cell.border = border
                    
                    # Appliquer des styles aux données (alternance de couleurs)
                    for row in range(3, ws.max_row + 1):
                        for col in range(1, last_col + 1):
                            cell = ws.cell(row=row, column=col)
                            cell.border = border
                            cell.alignment = Alignment(horizontal='center')
                            
                            # Lignes alternées
                            if row % 2 == 0:  # Lignes paires
                                cell.fill = alt_row_fill
                    
                    # Ajuster la largeur des colonnes
                    for col in range(1, last_col + 1):
                        column_letter = get_column_letter(col)
                        max_length = 0
                        for row in range(1, ws.max_row + 1):
                            cell = ws.cell(row=row, column=col)
                            if cell.value:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                        adjusted_width = max_length + 2
                        ws.column_dimensions[column_letter].width = min(adjusted_width, 25)
                
                # Ajouter une feuille de graphiques si possible
                try:
                    # Créer une nouvelle feuille pour les graphiques
                    charts_sheet = workbook.create_sheet("Graphiques")
                    
                    # Ajouter un titre explicatif sur comment voir les graphiques
                    charts_sheet.merge_cells('A1:J1')
                    charts_sheet['A1'] = "GRAPHIQUES DISPONIBLES DANS L'APPLICATION"
                    charts_sheet['A1'].font = title_font
                    charts_sheet['A1'].fill = title_fill
                    charts_sheet['A1'].alignment = center_align
                    
                    charts_sheet.merge_cells('A3:J3')
                    charts_sheet['A3'] = "Pour visualiser les graphiques des simulations, utilisez l'onglet 'Graphiques' dans l'application."
                    charts_sheet['A3'].font = Font(name='Arial', size=12, bold=True)
                    charts_sheet['A3'].alignment = Alignment(horizontal='center')
                    
                    # Instructions
                    charts_sheet.merge_cells('A5:J5')
                    charts_sheet['A5'] = "Les données temporelles complètes sont disponibles dans les onglets individuels de chaque simulation."
                    charts_sheet['A5'].font = Font(name='Arial', size=11)
                    charts_sheet['A5'].alignment = Alignment(horizontal='center')
                    
                    charts_sheet.row_dimensions[1].height = 30
                except:
                    pass  # En cas d'erreur, ignorer simplement cette partie
            
            messagebox.showinfo("Succès", f"Données exportées avec succès vers {os.path.basename(file_path)} avec mise en forme améliorée!")
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'exportation des données: {str(e)}")
    
    def show_data_summary(self):
        self.update_data_display()
    
    def create_simulation_tab(self):
        """Create a tab for setting simulation parameters"""
        container = ctk.CTkFrame(self.simulation_frame)
        container.pack(expand=True, fill="both", padx=15, pady=15)
        
        # Title with background
        title_frame = ctk.CTkFrame(container, fg_color="#3a7ebf")
        title_frame.pack(fill="x", padx=10, pady=(0, 15))
        
        title_label = ctk.CTkLabel(
            title_frame, 
            text="Paramètres de Simulation", 
            font=("Arial", 18, "bold"),
            text_color="white"
        )
        title_label.pack(pady=10)
        
        # Use a grid layout for better organization
        params_grid = ctk.CTkFrame(container)
        params_grid.pack(fill="both", expand=True, padx=10, pady=10)
        params_grid.grid_columnconfigure(0, weight=1)
        params_grid.grid_columnconfigure(1, weight=1)
        
        # === Left Column - Basic Parameters ===
        left_frame = ctk.CTkFrame(params_grid)
        left_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        
        # Section title
        ctk.CTkLabel(
            left_frame, 
            text="Paramètres de base", 
            font=("Arial", 14, "bold")
        ).pack(anchor="w", padx=10, pady=(10, 15))
        
        # HRT (Hydraulic Retention Time)
        param_frame = ctk.CTkFrame(left_frame)
        param_frame.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(
            param_frame, 
            text="Temps de rétention hydraulique (HRT):",
            anchor="w",
            width=200
        ).pack(side="left", padx=10, pady=5)
        
        hrt_container = ctk.CTkFrame(param_frame)
        hrt_container.pack(side="right", fill="x", expand=True, padx=10, pady=5)
        
        self.hrt_var = ctk.StringVar(value="840")
        self.hrt_entry = ctk.CTkEntry(hrt_container, textvariable=self.hrt_var, width=100)
        self.hrt_entry.pack(side="left", padx=5)
        
        ctk.CTkLabel(hrt_container, text="heures").pack(side="left")
        
        # Air flow (Débit d'air)
        param_frame = ctk.CTkFrame(left_frame)
        param_frame.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(
            param_frame, 
            text="Débit d'air:",
            anchor="w",
            width=200
        ).pack(side="left", padx=10, pady=5)
        
        air_flow_container = ctk.CTkFrame(param_frame)
        air_flow_container.pack(side="right", fill="x", expand=True, padx=10, pady=5)
        
        self.air_flow_var = ctk.StringVar(value="10.0")
        self.air_flow_entry = ctk.CTkEntry(air_flow_container, textvariable=self.air_flow_var, width=100)
        self.air_flow_entry.pack(side="left", padx=5)
        
        ctk.CTkLabel(air_flow_container, text="m³/h").pack(side="left")
        
        # Ambient temperature (Température ambiante)
        param_frame = ctk.CTkFrame(left_frame)
        param_frame.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(
            param_frame, 
            text="Température ambiante:",
            anchor="w",
            width=200
        ).pack(side="left", padx=10, pady=5)
        
        temp_container = ctk.CTkFrame(param_frame)
        temp_container.pack(side="right", fill="x", expand=True, padx=10, pady=5)
        
        self.temp_var = ctk.StringVar(value="25.0")
        self.temp_entry = ctk.CTkEntry(temp_container, textvariable=self.temp_var, width=100)
        self.temp_entry.pack(side="left", padx=5)
        
        ctk.CTkLabel(temp_container, text="°C").pack(side="left")
        
        # Relative humidity (Humidité relative)
        param_frame = ctk.CTkFrame(left_frame)
        param_frame.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(
            param_frame, 
            text="Humidité relative de l'air:",
            anchor="w",
            width=200
        ).pack(side="left", padx=10, pady=5)
        
        rh_container = ctk.CTkFrame(param_frame)
        rh_container.pack(side="right", fill="x", expand=True, padx=10, pady=5)
        
        self.rh_var = ctk.StringVar(value="60.0")
        self.rh_entry = ctk.CTkEntry(rh_container, textvariable=self.rh_var, width=100)
        self.rh_entry.pack(side="left", padx=5)
        
        ctk.CTkLabel(rh_container, text="%").pack(side="left")
        
        # === Right Column - Advanced Parameters ===
        right_frame = ctk.CTkFrame(params_grid)
        right_frame.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)
        
        # Section title
        ctk.CTkLabel(
            right_frame, 
            text="Paramètres avancés", 
            font=("Arial", 14, "bold")
        ).pack(anchor="w", padx=10, pady=(10, 15))
        
        # Air alternance
        self.air_alt_frame = ctk.CTkFrame(right_frame)
        self.air_alt_frame.pack(fill="x", padx=10, pady=10)
        
        # Option to enable alternance with better styling
        air_alt_title = ctk.CTkFrame(self.air_alt_frame)
        air_alt_title.pack(fill="x", padx=5, pady=5)
        
        self.air_alt_var = ctk.BooleanVar(value=False)
        air_alt_checkbox = ctk.CTkCheckBox(
            air_alt_title, 
            text="Activer l'alternance du débit d'air (On/Off)", 
            variable=self.air_alt_var, 
            command=self.toggle_air_alternance,
            font=("Arial", 12, "bold")
        )
        air_alt_checkbox.pack(side="left", padx=10, pady=5)
        
        # Frame for alternance parameters (initially hidden)
        self.air_alt_params_frame = ctk.CTkFrame(right_frame, fg_color="#f0f0f0")
        
        # ON duration
        on_frame = ctk.CTkFrame(self.air_alt_params_frame)
        on_frame.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(on_frame, text="Durée ON:", width=100).pack(side="left", padx=10)
        
        on_container = ctk.CTkFrame(on_frame)
        on_container.pack(side="right", fill="x", expand=True, padx=10, pady=5)
        
        self.air_on_var = ctk.StringVar(value="1.0")
        self.air_on_entry = ctk.CTkEntry(on_container, textvariable=self.air_on_var, width=100)
        self.air_on_entry.pack(side="left", padx=5)
        
        ctk.CTkLabel(on_container, text="heures").pack(side="left")
        
        # OFF duration
        off_frame = ctk.CTkFrame(self.air_alt_params_frame)
        off_frame.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(off_frame, text="Durée OFF:", width=100).pack(side="left", padx=10)
        
        off_container = ctk.CTkFrame(off_frame)
        off_container.pack(side="right", fill="x", expand=True, padx=10, pady=5)
        
        self.air_off_var = ctk.StringVar(value="0.5")
        self.air_off_entry = ctk.CTkEntry(off_container, textvariable=self.air_off_var, width=100)
        self.air_off_entry.pack(side="left", padx=5)
        
        ctk.CTkLabel(off_container, text="heures").pack(side="left")
        
        # Water addition section
        water_section = ctk.CTkFrame(right_frame)
        water_section.pack(fill="x", padx=10, pady=(20, 10))
        
        ctk.CTkLabel(
            water_section, 
            text="Ajout d'eau", 
            font=("Arial", 12, "bold")
        ).pack(anchor="w", padx=10, pady=5)
        
        # Water flow
        water_frame = ctk.CTkFrame(water_section)
        water_frame.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(water_frame, text="Débit d'eau:", width=100).pack(side="left", padx=10)
        
        water_flow_container = ctk.CTkFrame(water_frame)
        water_flow_container.pack(side="right", fill="x", expand=True, padx=10, pady=5)
        
        self.water_flow_var = ctk.StringVar(value="1.0")
        self.water_flow_entry = ctk.CTkEntry(water_flow_container, textvariable=self.water_flow_var, width=100)
        self.water_flow_entry.pack(side="left", padx=5)
        
        ctk.CTkLabel(water_flow_container, text="kg/h").pack(side="left")
        
        # Water temperature
        water_temp_frame = ctk.CTkFrame(water_section)
        water_temp_frame.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(water_temp_frame, text="Température:", width=100).pack(side="left", padx=10)
        
        water_temp_container = ctk.CTkFrame(water_temp_frame)
        water_temp_container.pack(side="right", fill="x", expand=True, padx=10, pady=5)
        
        self.water_temp_var = ctk.StringVar(value="20.0")
        self.water_temp_entry = ctk.CTkEntry(water_temp_container, textvariable=self.water_temp_var, width=100)
        self.water_temp_entry.pack(side="left", padx=5)
        
        ctk.CTkLabel(water_temp_container, text="°C").pack(side="left")
        
        # Button to save parameters
        buttons_frame = ctk.CTkFrame(container)
        buttons_frame.pack(pady=20)
        
        save_button = ctk.CTkButton(
            buttons_frame, 
            text="Enregistrer les paramètres", 
            command=self.save_simulation_params,
            width=200,
            height=40,
            font=("Arial", 14, "bold"),
            fg_color="#28a745",
            hover_color="#218838"
        )
        save_button.pack(pady=5)
        
        # Help text
        help_text = ctk.CTkLabel(
            buttons_frame,
            text="Ces paramètres seront utilisés lors de la prochaine simulation",
            font=("Arial", 10),
            text_color="gray"
        )
        help_text.pack(pady=5)
    
    def toggle_air_alternance(self):
        """Show or hide air flow alternance parameters"""
        if self.air_alt_var.get():
            self.air_alt_params_frame.pack(fill="x", padx=10, pady=5, after=self.air_alt_frame)
        else:
            self.air_alt_params_frame.pack_forget()
    
    def save_simulation_params(self):
        """Save simulation parameters"""
        try:
            # Retrieve and validate values
            # S'assurer que les valeurs sont bien numériques en nettoyant toute entrée non-numérique
            hrt_str = self.hrt_var.get().strip()
            # Nettoyer la chaîne en supprimant tout caractère non numérique après le point décimal
            if '.' in hrt_str:
                parts = hrt_str.split('.')
                hrt_str = parts[0] + '.' + ''.join(c for c in parts[1] if c.isdigit())
            else:
                # Garder seulement les chiffres
                hrt_str = ''.join(c for c in hrt_str if c.isdigit())
                
            hrt = float(hrt_str)
            
            # Appliquer la même validation pour les autres valeurs numériques
            air_flow_str = self.air_flow_var.get().strip()
            if '.' in air_flow_str:
                parts = air_flow_str.split('.')
                air_flow_str = parts[0] + '.' + ''.join(c for c in parts[1] if c.isdigit())
            else:
                air_flow_str = ''.join(c for c in air_flow_str if c.isdigit())
            air_flow = float(air_flow_str)
            
            rh_str = self.rh_var.get().strip()
            if '.' in rh_str:
                parts = rh_str.split('.')
                rh_str = parts[0] + '.' + ''.join(c for c in parts[1] if c.isdigit())
            else:
                rh_str = ''.join(c for c in rh_str if c.isdigit())
            rh = float(rh_str)
            
            temp_str = self.temp_var.get().strip()
            if '.' in temp_str:
                parts = temp_str.split('.')
                temp_str = parts[0] + '.' + ''.join(c for c in parts[1] if c.isdigit())
            else:
                temp_str = ''.join(c for c in temp_str if c.isdigit())
            temp = float(temp_str)
            
            water_flow_str = self.water_flow_var.get().strip()
            if '.' in water_flow_str:
                parts = water_flow_str.split('.')
                water_flow_str = parts[0] + '.' + ''.join(c for c in parts[1] if c.isdigit())
            else:
                water_flow_str = ''.join(c for c in water_flow_str if c.isdigit())
            water_flow = float(water_flow_str)
            
            water_temp_str = self.water_temp_var.get().strip()
            if '.' in water_temp_str:
                parts = water_temp_str.split('.')
                water_temp_str = parts[0] + '.' + ''.join(c for c in parts[1] if c.isdigit())
            else:
                water_temp_str = ''.join(c for c in water_temp_str if c.isdigit())
            water_temp = float(water_temp_str)
            
            # Alternance parameters
            air_alternance = self.air_alt_var.get()
            
            air_on_time_str = self.air_on_var.get().strip()
            if '.' in air_on_time_str:
                parts = air_on_time_str.split('.')
                air_on_time_str = parts[0] + '.' + ''.join(c for c in parts[1] if c.isdigit())
            else:
                air_on_time_str = ''.join(c for c in air_on_time_str if c.isdigit())
            air_on_time = float(air_on_time_str) if air_alternance else 0
            
            air_off_time_str = self.air_off_var.get().strip()
            if '.' in air_off_time_str:
                parts = air_off_time_str.split('.')
                air_off_time_str = parts[0] + '.' + ''.join(c for c in parts[1] if c.isdigit())
            else:
                air_off_time_str = ''.join(c for c in air_off_time_str if c.isdigit())
            air_off_time = float(air_off_time_str) if air_alternance else 0
            
            # Store parameters
            self.simulation_params = {
                "HRT": hrt,
                "air_flow": air_flow,
                "relative_humidity": rh,
                "ambient_temp": temp,
                "water_flow": water_flow,
                "water_temp": water_temp,
                "air_alternance": air_alternance,
                "air_on_time": air_on_time,
                "air_off_time": air_off_time
            }
            
            # Update data display
            self.update_data_display()
            
            messagebox.showinfo("Succès", "Paramètres de simulation enregistrés avec succès!")
        except ValueError as e:
            messagebox.showerror("Erreur", f"Veuillez entrer des valeurs numériques valides.\nErreur: {str(e)}")
    
    def update_data_display(self):
        """Update data display in the Data tab with properly formatted tables"""
        # Clear the current display
        self.data_text.delete("1.0", "end")
        
        # ===== SUBSTRATES TABLE =====
        substrates_title = "TABLEAU DES SUBSTRATS"
        self.data_text.insert("end", f"\n{substrates_title}\n")
        self.data_text.insert("end", "=" * 80 + "\n\n")
        
        # Chemical composition table
        chem_headers = ["Substrat", "a (C)", "b (H)", "c (O)", "d (N)"]
        chem_rows = []
        
        for sub in self.substrates:
            chem_rows.append([
                sub["name"], 
                sub["composition"][0], 
                sub["composition"][1],
                sub["composition"][2],
                sub["composition"][3]
            ])
        
        chem_table = self.create_formatted_table(
            chem_headers, 
            chem_rows, 
            "Composition chimique des substrats"
        )
        self.data_text.insert("end", chem_table + "\n")
        
        # Properties table
        prop_headers = ["Substrat", "T (°C)", "FR (kg/h)", "FS (%)", "FVS (%)", "FBVS (%)", "FfBVS (%)", "fKT20", "sKT20"]
        prop_rows = []
        
        for sub in self.substrates:
            prop_rows.append([
                sub["name"], 
                sub["T"],
                sub["FR"],
                sub["FS"],
                sub["FVS"],
                sub["FBVS"],
                sub["FfBVS"],
                sub["fKT20"],
                sub["sKT20"]
            ])
        
        prop_table = self.create_formatted_table(
            prop_headers, 
            prop_rows, 
            "Propriétés des substrats"
        )
        self.data_text.insert("end", prop_table + "\n")
        
        # ===== SIMULATION PARAMETERS TABLE =====
        if hasattr(self, 'simulation_params'):
            sim_title = "PARAMÈTRES DE SIMULATION ACTUELS"
            self.data_text.insert("end", f"\n{sim_title}\n")
            self.data_text.insert("end", "=" * 80 + "\n\n")
            
            sim_headers = ["Paramètre", "Valeur", "Unité"]
            sim_rows = [
                ["Temps de rétention hydraulique", self.simulation_params['HRT'], "heures"],
                ["Débit d'air", self.simulation_params['air_flow'], "m³/h"],
                ["Humidité relative de l'air", self.simulation_params['relative_humidity'], "%"],
                ["Température ambiante", self.simulation_params['ambient_temp'], "°C"],
                ["Débit d'eau ajouté", self.simulation_params['water_flow'], "kg/h"],
                ["Température de l'eau", self.simulation_params['water_temp'], "°C"],
            ]
            
            # Add alternance parameters if enabled
            if self.simulation_params['air_alternance']:
                sim_rows.append(["Alternance du débit d'air", "Activée", ""])
                sim_rows.append(["Durée ON", self.simulation_params['air_on_time'], "heures"])
                sim_rows.append(["Durée OFF", self.simulation_params['air_off_time'], "heures"])
            else:
                sim_rows.append(["Alternance du débit d'air", "Désactivée", ""])
            
            sim_table = self.create_formatted_table(sim_headers, sim_rows)
            self.data_text.insert("end", sim_table + "\n")
        
        # ===== SIMULATION RESULTS TABLES =====
        if hasattr(self, 'simulations') and self.simulations:
            sim_results_title = "SIMULATIONS EFFECTUÉES"
            self.data_text.insert("end", f"\n{sim_results_title}\n")
            self.data_text.insert("end", "=" * 80 + "\n\n")
            
            for i, sim in enumerate(self.simulations):
                # Display simulation name and number
                self.data_text.insert("end", f"Simulation #{i+1}: {sim['name']}\n\n")
                
                # Parameters table
                param_headers = ["Paramètre", "Valeur", "Unité"]
                param_rows = [
                    ["Temps de rétention hydraulique", sim['params']['HRT'], "heures"],
                    ["Débit d'air", sim['params']['air_flow'], "m³/h"],
                    ["Humidité relative de l'air", sim['params']['relative_humidity'], "%"],
                    ["Température ambiante", sim['params']['ambient_temp'], "°C"],
                    ["Débit d'eau ajouté", sim['params']['water_flow'], "kg/h"],
                    ["Température de l'eau", sim['params']['water_temp'], "°C"],
                ]
                
                # Add alternance parameters if enabled
                if sim['params']['air_alternance']:
                    param_rows.append(["Alternance du débit d'air", "Activée", ""])
                    param_rows.append(["Durée ON", sim['params']['air_on_time'], "heures"])
                    param_rows.append(["Durée OFF", sim['params']['air_off_time'], "heures"])
                else:
                    param_rows.append(["Alternance du débit d'air", "Désactivée", ""])
                
                param_table = self.create_formatted_table(
                    param_headers, 
                    param_rows, 
                    "Paramètres utilisés"
                )
                self.data_text.insert("end", param_table + "\n")
                
                # Results table
                result_headers = ["Paramètre", "Valeur initiale", "Valeur finale", "Unité"]
                result_rows = []
                
                # Extract initial and final values for key parameters
                if "Temperatures" in sim["data"] and len(sim["data"]["Temperatures"]) >= 2:
                    result_rows.append([
                        "Température", 
                        sim["data"]["Temperatures"][0], 
                        sim["data"]["Temperatures"][-1], 
                        "°C"
                    ])
                
                if "MoistureFraction" in sim["data"] and len(sim["data"]["MoistureFraction"]) >= 2:
                    result_rows.append([
                        "Fraction d'humidité", 
                        sim["data"]["MoistureFraction"][0], 
                        sim["data"]["MoistureFraction"][-1], 
                        "%"
                    ])
                
                if "Solids" in sim["data"] and len(sim["data"]["Solids"]) >= 2:
                    result_rows.append([
                        "Solides", 
                        sim["data"]["Solids"][0], 
                        sim["data"]["Solids"][-1], 
                        "kg"
                    ])
                
                if "QExhaustgases" in sim["data"] and len(sim["data"]["QExhaustgases"]) >= 2:
                    result_rows.append([
                        "Débit de gaz d'échappement", 
                        sim["data"]["QExhaustgases"][0], 
                        sim["data"]["QExhaustgases"][-1], 
                        "kg/h"
                    ])
                
                if "process_volume" in sim["data"]:
                    result_rows.append([
                        "Volume du processus", 
                        "-", 
                        sim["data"]["process_volume"], 
                        "m³"
                    ])
                
                result_table = self.create_formatted_table(
                    result_headers, 
                    result_rows, 
                    "Résultats de la simulation"
                )
                self.data_text.insert("end", result_table + "\n\n")
                
                # Add separator between simulations
                if i < len(self.simulations) - 1:
                    self.data_text.insert("end", "-" * 80 + "\n\n")
    
    def create_formatted_table(self, headers, rows, title=None):
        """Create a nicely formatted table with proper alignment
        
        Args:
            headers: List of column headers
            rows: List of rows (each row is a list of values)
            title: Optional title for the table
            
        Returns:
            Formatted table as string
        """
        # Determine column widths based on content
        col_widths = [len(str(h)) for h in headers]
        for row in rows:
            for i, cell in enumerate(row):
                col_widths[i] = max(col_widths[i], len(str(cell)))
        
        # Add padding to column widths
        col_widths = [w + 4 for w in col_widths]
        
        # Calculate total width of table
        total_width = sum(col_widths) + len(headers) + 1
        
        # Format title if provided
        result = ""
        if title:
            result += f"\n{title}\n"
            result += "=" * total_width + "\n"
        
        # Create header row
        header_row = "│"
        for i, header in enumerate(headers):
            header_row += f" {header:{col_widths[i]}} │"
        result += header_row + "\n"
        
        # Create separator line
        separator = "├"
        for width in col_widths:
            separator += "─" * (width + 2) + "┼"
        result = result[:-1] + "┐\n" + separator[:-1] + "┤\n"
        
        # Create data rows
        for row in rows:
            data_row = "│"
            for i, cell in enumerate(row):
                # Format numbers properly
                if isinstance(cell, (int, float)):
                    if isinstance(cell, int):
                        formatted_cell = f"{cell:,d}"
                    else:
                        formatted_cell = f"{cell:.2f}"
                else:
                    formatted_cell = str(cell)
                data_row += f" {formatted_cell:{col_widths[i]}} │"
            result += data_row + "\n"
        
        # Add bottom border
        bottom_border = "└"
        for width in col_widths:
            bottom_border += "─" * (width + 2) + "┴"
        result += bottom_border[:-1] + "┘\n"
        
        return result
    
    def create_graph_tab(self):
        # Frame principale pour l'onglet Graphiques
        graph_container = ctk.CTkFrame(self.graph_frame)
        graph_container.pack(expand=True, fill="both", padx=15, pady=15)
        
        # Panel de contrôle commun en haut
        controls_panel = ctk.CTkFrame(graph_container)
        controls_panel.pack(side="top", fill="x", padx=5, pady=5)
        
        # Utiliser le layout grid pour une meilleure organisation
        controls_panel.grid_columnconfigure(0, weight=1)  # Nom de simulation
        controls_panel.grid_columnconfigure(1, weight=2)  # Boutons d'action
        
        # Champ pour le nom de la simulation
        name_frame = ctk.CTkFrame(controls_panel)
        name_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
        
        ctk.CTkLabel(
            name_frame, 
            text="Nom de simulation:",
            font=("Arial", 12, "bold")
        ).pack(side="left", padx=5, pady=5)
        
        self.sim_name_var = ctk.StringVar(value="Simulation 1")
        sim_name_entry = ctk.CTkEntry(name_frame, textvariable=self.sim_name_var, width=150)
        sim_name_entry.pack(side="left", padx=5, pady=5)
        
        # Boutons d'action
        buttons_frame = ctk.CTkFrame(controls_panel)
        buttons_frame.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
        
        # Définir des méthodes qui seront utilisées comme callbacks
        def run_sim_callback():
            self.run_and_store_simulation()
            
        def plot_solid_callback():
            if hasattr(self, 'plot_solid_graphs'):
                self.plot_solid_graphs()
            else:
                messagebox.showwarning("Avertissement", "Fonction d'affichage non disponible pour la phase solide.")
                
        def plot_liquid_callback():
            if hasattr(self, 'plot_liquid_graphs'):
                self.plot_liquid_graphs()
            else:
                messagebox.showwarning("Avertissement", "Fonction d'affichage non disponible pour la phase liquide.")
                
        def plot_gas_callback():
            if hasattr(self, 'plot_gas_graphs'):
                self.plot_gas_graphs()
            else:
                messagebox.showwarning("Avertissement", "Fonction d'affichage non disponible pour la phase gazeuse.")
            
        def plot_callback():
            active_tab = self.phase_tabs.get()
            if active_tab == "Solide":
                plot_solid_callback()
            elif active_tab == "Liquide":
                plot_liquid_callback()
            elif active_tab == "Gazeuse":
                plot_gas_callback()
                
        def export_callback():
            self.export_current_graph()
            
        def clear_callback():
            self.clear_simulations()
        
        # Bouton Exécuter
        run_btn = ctk.CTkButton(
            buttons_frame,
            text="Lancer",
            command=run_sim_callback,
            fg_color="#28a745",
            hover_color="#218838",
            width=90,
            height=30
        )
        run_btn.pack(side="left", padx=5, pady=5)
        
        # Bouton Afficher
        plot_btn = ctk.CTkButton(
            buttons_frame,
            text="Afficher",
            command=plot_callback,
            width=90,
            height=30
        )
        plot_btn.pack(side="left", padx=5, pady=5)
        
        # Bouton Exporter
        export_btn = ctk.CTkButton(
            buttons_frame,
            text="Exporter",
            command=export_callback,
            width=90,
            height=30
        )
        export_btn.pack(side="left", padx=5, pady=5)
        
        # Bouton Effacer
        clear_btn = ctk.CTkButton(
            buttons_frame,
            text="Effacer",
            command=clear_callback,
            fg_color="#dc3545",
            hover_color="#c82333",
            width=90,
            height=30
        )
        clear_btn.pack(side="left", padx=5, pady=5)
        
        # Stocker les résultats de simulation
        self.simulations = []
        self.simulation_names = []
        
        # Créer la zone principale avec les onglets par phase et la liste des simulations
        main_area = ctk.CTkFrame(graph_container)
        main_area.pack(side="bottom", fill="both", expand=True, padx=5, pady=5)
        
        # Division: graphiques (85%) et liste des simulations (15%)
        main_area.grid_columnconfigure(0, weight=85)  # Zone graphique
        main_area.grid_rowconfigure(0, weight=100)    # Hauteur complète
        
        # Zone des graphiques avec onglets pour chaque phase
        graph_tab_frame = ctk.CTkFrame(main_area)
        graph_tab_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        
        # Créer les onglets pour les phases physiques
        self.phase_tabs = ctk.CTkTabview(graph_tab_frame)
        self.phase_tabs.pack(fill="both", expand=True)
        
        # Ajouter les onglets pour chaque phase
        self.solid_tab = self.phase_tabs.add("Solide")
        self.liquid_tab = self.phase_tabs.add("Liquide")
        self.gas_tab = self.phase_tabs.add("Gazeuse")
        
        # Configuration pour l'onglet Solide
        self.setup_solid_tab()
        
        # Configuration pour l'onglet Liquide
        self.setup_liquid_tab()
        
        # Configuration pour l'onglet Gazeuse
        self.setup_gas_tab()
        
        # Zone d'affichage des simulations actives
        sim_display_frame = ctk.CTkFrame(main_area)
        sim_display_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
        
        ctk.CTkLabel(
            sim_display_frame, 
            text="Simulations actives:", 
            font=("Arial", 12, "bold")
        ).pack(anchor="w", padx=5, pady=5)
        
        self.sim_display = ctk.CTkTextbox(sim_display_frame, wrap="word", width=150)
        self.sim_display.pack(fill="both", expand=True, padx=5, pady=5)
    
    def setup_solid_tab(self):
        """Configurer l'onglet pour la phase solide"""
        container = ctk.CTkFrame(self.solid_tab)
        container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # En-tête avec sélection des courbes
        header = ctk.CTkFrame(container)
        header.pack(fill="x", padx=5, pady=5)
        
        ctk.CTkLabel(
            header,
            text="Courbes de la phase solide:",
            font=("Arial", 12, "bold")
        ).pack(side="left", padx=10, pady=5)
        
        # Options des courbes pour la phase solide
        solid_options = [
            "Matière sèche (MS)",
            "Température",
            "Rapport C/N"
        ]
        
        # Création des checkboxes pour la sélection multiple
        self.solid_vars = []
        options_frame = ctk.CTkFrame(header)
        options_frame.pack(side="left", fill="x", expand=True, padx=10, pady=5)
        
        for i, option in enumerate(solid_options):
            var = ctk.BooleanVar(value=(i == 0))  # Premier sélectionné par défaut
            checkbox = ctk.CTkCheckBox(
                options_frame,
                text=option,
                variable=var
            )
            checkbox.pack(side="left", padx=10)
            self.solid_vars.append((var, option))
        
        # Zone du graphique
        self.solid_fig_frame = ctk.CTkFrame(container)
        self.solid_fig_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.solid_figure = plt.Figure(figsize=(10, 6), dpi=100)
        self.solid_canvas = FigureCanvasTkAgg(self.solid_figure, self.solid_fig_frame)
        self.solid_canvas.get_tk_widget().pack(fill="both", expand=True)
        
        # Ajouter une barre d'outils de navigation
        toolbar_frame = ctk.CTkFrame(self.solid_fig_frame)
        toolbar_frame.pack(side="bottom", fill="x")
        toolbar = NavigationToolbar2Tk(self.solid_canvas, toolbar_frame)
        toolbar.update()
    
    def setup_liquid_tab(self):
        """Configurer l'onglet pour la phase liquide"""
        container = ctk.CTkFrame(self.liquid_tab)
        container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # En-tête avec sélection des courbes
        header = ctk.CTkFrame(container)
        header.pack(fill="x", padx=5, pady=5)
        
        ctk.CTkLabel(
            header,
            text="Courbes de la phase liquide:",
            font=("Arial", 12, "bold")
        ).pack(side="left", padx=10, pady=5)
        
        # Options des courbes pour la phase liquide
        liquid_options = [
            "Humidité massique",
            "Humidité relative"
        ]
        
        # Création des checkboxes pour la sélection multiple
        self.liquid_vars = []
        options_frame = ctk.CTkFrame(header)
        options_frame.pack(side="left", fill="x", expand=True, padx=10, pady=5)
        
        for i, option in enumerate(liquid_options):
            var = ctk.BooleanVar(value=(i == 0))  # Premier sélectionné par défaut
            checkbox = ctk.CTkCheckBox(
                options_frame,
                text=option,
                variable=var
            )
            checkbox.pack(side="left", padx=10)
            self.liquid_vars.append((var, option))
        
        # Zone du graphique
        self.liquid_fig_frame = ctk.CTkFrame(container)
        self.liquid_fig_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.liquid_figure = plt.Figure(figsize=(10, 6), dpi=100)
        self.liquid_canvas = FigureCanvasTkAgg(self.liquid_figure, self.liquid_fig_frame)
        self.liquid_canvas.get_tk_widget().pack(fill="both", expand=True)
        
        # Ajouter une barre d'outils de navigation
        toolbar_frame = ctk.CTkFrame(self.liquid_fig_frame)
        toolbar_frame.pack(side="bottom", fill="x")
        toolbar = NavigationToolbar2Tk(self.liquid_canvas, toolbar_frame)
        toolbar.update()
    
    def setup_gas_tab(self):
        """Configurer l'onglet pour la phase gazeuse"""
        container = ctk.CTkFrame(self.gas_tab)
        container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # En-tête avec sélection des courbes
        header = ctk.CTkFrame(container)
        header.pack(fill="x", padx=5, pady=5)
        
        ctk.CTkLabel(
            header,
            text="Courbes de la phase gazeuse:",
            font=("Arial", 12, "bold")
        ).pack(side="left", padx=10, pady=5)
        
        # Options des courbes pour la phase gazeuse
        gas_options = [
            "CO₂ généré",
            "O₂ consommé",
            "NH₃ émis",
            "Débit de gaz d'échappement",
            "Volume de gaz d'échappement"
        ]
        
        # Création des checkboxes pour la sélection multiple
        self.gas_vars = []
        options_frame = ctk.CTkFrame(header)
        options_frame.pack(side="left", fill="x", expand=True, padx=10, pady=5)
        
        for i, option in enumerate(gas_options):
            var = ctk.BooleanVar(value=(i == 0))  # Premier sélectionné par défaut
            checkbox = ctk.CTkCheckBox(
                options_frame,
                text=option,
                variable=var
            )
            checkbox.pack(side="left", padx=10)
            self.gas_vars.append((var, option))
        
        # Zone du graphique
        self.gas_fig_frame = ctk.CTkFrame(container)
        self.gas_fig_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.gas_figure = plt.Figure(figsize=(10, 6), dpi=100)
        self.gas_canvas = FigureCanvasTkAgg(self.gas_figure, self.gas_fig_frame)
        self.gas_canvas.get_tk_widget().pack(fill="both", expand=True)
        
        # Ajouter une barre d'outils de navigation
        toolbar_frame = ctk.CTkFrame(self.gas_fig_frame)
        toolbar_frame.pack(side="bottom", fill="x")
        toolbar = NavigationToolbar2Tk(self.gas_canvas, toolbar_frame)
        toolbar.update()
    
    def test_solid_graph(self):
        """Afficher un graphique de test simple"""
        # Effacer le graphique actuel
        self.solid_figure.clear()
        
        # Créer des données de test
        x = np.linspace(0, 10, 100)
        y = np.sin(x)
        
        # Créer un subplot et tracer les données
        ax = self.solid_figure.add_subplot(111)
        ax.plot(x, y, 'b-', linewidth=2, label='Test')
        
        # Configurer le graphique
        ax.set_xlabel("Temps (h)", fontsize=12)
        ax.set_ylabel("Valeur", fontsize=12)
        ax.set_title("Graphique de test", fontsize=14, fontweight='bold')
        ax.grid(True, linestyle='--', alpha=0.7)
        ax.legend()
        
        # Ajuster la mise en page
        self.solid_figure.tight_layout()
        
        # Mettre à jour le canvas
        self.solid_canvas.draw()
        
        # Afficher un message de diagnostic
        print("Test de graphique exécuté. Le graphique devrait être visible dans l'onglet Solide.")
    
    def test_liquid_graph(self):
        """Afficher un graphique de test simple"""
        # Effacer le graphique actuel
        self.liquid_figure.clear()
        
        # Créer des données de test
        x = np.linspace(0, 10, 100)
        y = np.cos(x)
        
        # Créer un subplot et tracer les données
        ax = self.liquid_figure.add_subplot(111)
        ax.plot(x, y, 'g-', linewidth=2, label='Test')
        
        # Configurer le graphique
        ax.set_xlabel("Temps (h)", fontsize=12)
        ax.set_ylabel("Valeur", fontsize=12)
        ax.set_title("Graphique de test - Phase liquide", fontsize=14, fontweight='bold')
        ax.grid(True, linestyle='--', alpha=0.7)
        ax.legend()
        
        # Ajuster la mise en page
        self.liquid_figure.tight_layout()
        
        # Mettre à jour le canvas
        self.liquid_canvas.draw()
        
        # Afficher un message de diagnostic
        print("Test de graphique liquide exécuté. Le graphique devrait être visible dans l'onglet Liquide.")
    
    def test_gas_graph(self):
        """Afficher un graphique de test simple pour l'onglet gaz"""
        # Effacer le graphique actuel
        self.gas_figure.clear()
        
        # Créer des données de test
        x = np.linspace(0, 10, 100)
        y1 = np.exp(-x/5) * np.sin(x)
        y2 = np.exp(-x/5) * np.cos(x)
        
        # Créer un subplot et tracer les données
        ax = self.gas_figure.add_subplot(111)
        ax.plot(x, y1, 'r-', linewidth=2, label='Test 1')
        ax.plot(x, y2, 'b--', linewidth=2, label='Test 2')
        
        # Configurer le graphique
        ax.set_xlabel("Temps (h)", fontsize=12)
        ax.set_ylabel("Valeur", fontsize=12)
        ax.set_title("Graphique de test - Phase gazeuse", fontsize=14, fontweight='bold')
        ax.grid(True, linestyle='--', alpha=0.7)
        ax.legend()
        
        # Ajuster la mise en page
        self.gas_figure.tight_layout()
        
        # Mettre à jour le canvas
        self.gas_canvas.draw()
        
        # Afficher un message de diagnostic
        print("Test de graphique gaz exécuté. Le graphique devrait être visible dans l'onglet Gazeuse.")
    
    def calculate_gas_emissions(self, data):
        """Calcule les émissions de gaz (CO2, O2, NH3) basées sur la dégradation de la matière organique"""
        # Vérifier si les données nécessaires sont disponibles
        if not all(key in data for key in ["Times", "Solids"]):
            return {}, False
            
        # Récupérer les données de dégradation
        times = data["Times"]
        initial_solids = data["Solids"][0]
        final_solids = data["Solids"][-1]
        solids = data["Solids"]
        
        # Initialiser les tableaux pour stocker les valeurs calculées
        co2_values = []
        o2_values = []
        nh3_values = []
        
        # Facteurs stœchiométriques approximatifs pour la dégradation aérobie
        # Basés sur la composition typique de la matière organique C₂₇H₃₈O₁₆N
        CO2_FACTOR = 1.5  # kg CO2 / kg matière organique dégradée
        O2_FACTOR = 1.2   # kg O2 / kg matière organique dégradée
        NH3_FACTOR = 0.05 # kg NH3 / kg matière organique dégradée
        
        # Calcul cumulatif des émissions
        cumulative_co2 = 0
        cumulative_o2 = 0
        cumulative_nh3 = 0
        
        for i in range(len(times)):
            if i > 0:
                # Calcul de la matière dégradée entre ce pas de temps et le précédent
                degraded = solids[i-1] - solids[i]
                if degraded > 0:
                    # Conversion de la matière dégradée en gaz émis
                    co2_produced = degraded * CO2_FACTOR
                    o2_consumed = degraded * O2_FACTOR
                    nh3_produced = degraded * NH3_FACTOR
                    
                    # Accumulation
                    cumulative_co2 += co2_produced
                    cumulative_o2 += o2_consumed
                    cumulative_nh3 += nh3_produced
            
            # Stocker les valeurs cumulatives
            co2_values.append(cumulative_co2)
            o2_values.append(cumulative_o2)
            nh3_values.append(cumulative_nh3)
        
        # Retourner les résultats calculés
        gas_data = {
            "CO2": co2_values,
            "O2": o2_values,
            "NH3": nh3_values
        }
        
        return gas_data, True

    def plot_gas_graphs(self):
        """Afficher les graphiques de la phase gazeuse"""
        # Effacer le graphique actuel
        self.gas_figure.clear()
        
        # Récupérer les options sélectionnées
        selected_options = [option for var, option in self.gas_vars if var.get()]
        
        if not selected_options:
            messagebox.showinfo("Information", "Veuillez sélectionner au moins une courbe à afficher.")
            return
        
        # Créer un subplot unique
        ax = self.gas_figure.add_subplot(111)
        
        # Palette de couleurs pour différentes simulations
        colors = plt.cm.tab10.colors
        
        # Stocker les données pour pouvoir les analyser
        co2_data = {}
        o2_data = {}
        nh3_data = {}
        flow_data = {}
        volume_data = {}
        
        # Tracer les courbes sélectionnées pour chaque simulation
        legend_added = False
        for j, sim in enumerate(self.simulations):
            data = sim["data"]
            sim_name = sim["name"]
            color = colors[j % len(colors)]
            
            # Calculer les émissions de gaz si nécessaire
            gas_data = {}
            if any(option in ["CO₂ généré", "O₂ consommé", "NH₃ émis"] for option in selected_options):
                gas_data, success = self.calculate_gas_emissions(data)
            
            for option in selected_options:
                if option == "CO₂ généré" and "CO2" in gas_data:
                    ax.plot(data["Times"], gas_data["CO2"], 
                           label=f"{sim_name} - CO₂", color=color, linestyle='-', linewidth=2)
                    legend_added = True
                    co2_data[sim_name] = {
                        "times": data["Times"],
                        "values": gas_data["CO2"],
                        "color": color
                    }
                elif option == "O₂ consommé" and "O2" in gas_data:
                    ax.plot(data["Times"], gas_data["O2"], 
                           label=f"{sim_name} - O₂", color=color, linestyle='--', linewidth=2)
                    legend_added = True
                    o2_data[sim_name] = {
                        "times": data["Times"],
                        "values": gas_data["O2"],
                        "color": color
                    }
                elif option == "NH₃ émis" and "NH3" in gas_data:
                    ax.plot(data["Times"], gas_data["NH3"], 
                           label=f"{sim_name} - NH₃", color=color, linestyle=':', linewidth=2)
                    legend_added = True
                    nh3_data[sim_name] = {
                        "times": data["Times"],
                        "values": gas_data["NH3"],
                        "color": color
                    }
                elif option == "Débit de gaz d'échappement" and "QExhaustgases" in data:
                    ax.plot(data["Times"], data["QExhaustgases"], 
                           label=f"{sim_name} - Débit de gaz", color=color, linestyle='-.', linewidth=2)
                    legend_added = True
                    flow_data[sim_name] = {
                        "times": data["Times"],
                        "values": data["QExhaustgases"],
                        "color": color
                    }
                elif option == "Volume de gaz d'échappement" and "VExhaustgases" in data:
                    ax.plot(data["Times"], data["VExhaustgases"], 
                           label=f"{sim_name} - Volume de gaz", color=color, linestyle='--', linewidth=1.5)
                    legend_added = True
                    volume_data[sim_name] = {
                        "times": data["Times"],
                        "values": data["VExhaustgases"],
                        "color": color
                    }
        
        # Ajouter des annotations pour expliquer les phases d'activité
        # Pour cette analyse, nous utilisons principalement CO2 ou le débit de gaz
        analysis_data = None
        analysis_type = None
        
        if co2_data:
            # Utiliser CO2 pour analyser les phases d'activité biologique
            first_sim = list(co2_data.keys())[0]
            analysis_data = co2_data[first_sim]
            analysis_type = "CO2"
        elif flow_data:
            # Utiliser le débit de gaz comme indicateur d'activité
            first_sim = list(flow_data.keys())[0]
            analysis_data = flow_data[first_sim]
            analysis_type = "débit"
        
        if analysis_data:
            times = analysis_data["times"]
            values = analysis_data["values"]
            
            # Calculer la dérivée pour identifier les phases d'activité
            try:
                from scipy.signal import savgol_filter
                
                # Calculer les taux de changement
                rates = [0]
                for i in range(1, len(values)):
                    rate = (values[i] - values[i-1]) / (times[i] - times[i-1])
                    rates.append(rate)
                
                # Lisser les taux pour faciliter l'analyse
                smooth_window = min(11, len(rates) - 1 if len(rates) % 2 == 0 else len(rates))
                if smooth_window > 3:
                    smooth_rates = savgol_filter(rates, smooth_window, 3)
                else:
                    smooth_rates = rates
                
                # Identifier les phases clés
                # Phase 1: Lag phase (démarrage lent)
                # Phase 2: Accélération (activité croissante)
                # Phase 3: Activité maximale
                # Phase 4: Ralentissement
                # Phase 5: Stabilisation
                
                # Trouver les points d'inflexion dans les taux de changement
                inflection_points = []
                for i in range(1, len(smooth_rates)-1):
                    if (smooth_rates[i-1] < smooth_rates[i] and smooth_rates[i] > smooth_rates[i+1]) or \
                       (smooth_rates[i-1] > smooth_rates[i] and smooth_rates[i] < smooth_rates[i+1]):
                        inflection_points.append(i)
                
                # Si trop peu de points, utiliser une approche simplifiée
                if len(inflection_points) < 2:
                    # Division simplifiée en 3 phases
                    lag_end = len(times) // 5
                    active_end = 4 * len(times) // 5
                    stabilization_end = len(times) - 1
                else:
                    # Utiliser les points d'inflexion pour déterminer les phases
                    lag_end = inflection_points[0]
                    
                    # Trouver le point d'activité maximale
                    max_activity = max(values)
                    max_activity_idx = values.index(max_activity)
                    
                    # La phase active se termine lorsque l'activité commence à diminuer significativement
                    active_end = max_activity_idx
                    for i in inflection_points:
                        if i > max_activity_idx:
                            active_end = i
                            break
                    
                    # Le reste est la phase de stabilisation
                    stabilization_end = len(times) - 1
                
                # Colorer les phases identifiées
                # Phase de lag
                if lag_end > 0:
                    ax.axvspan(times[0], times[lag_end], color='blue', alpha=0.1)
                    ax.annotate('Phase de démarrage\n(activité initiale)', 
                             xy=((times[0] + times[lag_end])/2, values[lag_end//2]),
                             xytext=(0, 30),
                             textcoords='offset points',
                             ha='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="blue", alpha=0.7))
                
                # Phase active
                if active_end > lag_end:
                    ax.axvspan(times[lag_end], times[active_end], color='red', alpha=0.1)
                    ax.annotate('Phase d\'activité intense\n(production maximale de gaz)', 
                             xy=((times[lag_end] + times[active_end])/2, values[(lag_end+active_end)//2]),
                             xytext=(0, -30),
                             textcoords='offset points',
                             ha='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="red", alpha=0.7))
                
                # Phase de stabilisation
                if stabilization_end > active_end:
                    ax.axvspan(times[active_end], times[stabilization_end], color='green', alpha=0.1)
                    ax.annotate('Phase de stabilisation\n(ralentissement de l\'activité)', 
                             xy=((times[active_end] + times[stabilization_end])/2, values[(active_end+stabilization_end)//2]),
                             xytext=(0, 30),
                             textcoords='offset points',
                             ha='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="green", alpha=0.7))
                
                # Ajouter une légende pour les phases
                ax.legend(
                    handles=[
                        mpatches.Patch(color='blue', alpha=0.1, label='Phase de démarrage'),
                        mpatches.Patch(color='red', alpha=0.1, label='Phase d\'activité intense'),
                        mpatches.Patch(color='green', alpha=0.1, label='Phase de stabilisation')
                    ],
                    title="Phases d'activité",
                    loc="upper right",
                    fontsize=9
                )
                
            except Exception as e:
                print(f"Erreur lors de l'analyse des phases: {str(e)}")
        
        # Ajouter des annotations supplémentaires si nous avons des données d'ammoniac
        if nh3_data and analysis_data:
            # L'émission d'ammoniac est souvent liée à des périodes de haute température
            # et à un pH élevé. Annoter les pics d'émission d'ammoniac.
            first_sim = list(nh3_data.keys())[0]
            nh3_times = nh3_data[first_sim]["times"]
            nh3_values = nh3_data[first_sim]["values"]
            
            # Calculer la dérivée pour identifier les pics d'émission
            nh3_rates = [0]
            for i in range(1, len(nh3_values)):
                rate = (nh3_values[i] - nh3_values[i-1]) / (nh3_times[i] - nh3_times[i-1])
                nh3_rates.append(rate)
            
            # Identifier les pics d'émission d'ammoniac (taux de changement élevé)
            max_rate = max(nh3_rates)
            threshold = max_rate * 0.7  # 70% du taux maximal
            
            for i in range(1, len(nh3_rates)):
                if nh3_rates[i] > threshold:
                    ax.annotate('Pic d\'émission NH₃\nPossible pH élevé', 
                             xy=(nh3_times[i], nh3_values[i]),
                             xytext=(10, 10),
                             textcoords='offset points',
                             arrowprops=dict(arrowstyle="->", color="purple"),
                             bbox=dict(boxstyle="round,pad=0.2", fc="white", ec="purple", alpha=0.7),
                             fontsize=9)
                    break  # Une seule annotation pour éviter l'encombrement
        
        # Configurer le graphique
        ax.set_xlabel("Temps (h)", fontsize=12)
        
        # Définir l'unité appropriée en fonction des courbes sélectionnées
        if len(selected_options) == 1:
            if selected_options[0] == "CO₂ généré" or selected_options[0] == "O₂ consommé" or selected_options[0] == "NH₃ émis":
                ax.set_ylabel("Quantité (kg)", fontsize=12)
            elif selected_options[0] == "Débit de gaz d'échappement":
                ax.set_ylabel("Débit (kg/h)", fontsize=12)
            elif selected_options[0] == "Volume de gaz d'échappement":
                ax.set_ylabel("Volume (m³)", fontsize=12)
        else:
            # Si plusieurs courbes sont sélectionnées avec des unités différentes
            gas_emission = any(opt in ["CO₂ généré", "O₂ consommé", "NH₃ émis"] for opt in selected_options)
            flow_rate = "Débit de gaz d'échappement" in selected_options
            volume = "Volume de gaz d'échappement" in selected_options
            
            if gas_emission and not flow_rate and not volume:
                ax.set_ylabel("Quantité (kg)", fontsize=12)
            elif flow_rate and not gas_emission and not volume:
                ax.set_ylabel("Débit (kg/h)", fontsize=12)
            elif volume and not gas_emission and not flow_rate:
                ax.set_ylabel("Volume (m³)", fontsize=12)
            else:
                # Plusieurs types de données avec différentes unités
                ax.set_ylabel("Valeur (voir légende pour unités)", fontsize=12)
                
        ax.set_title("Évolution des paramètres de la phase gazeuse", fontsize=14, fontweight='bold')
        ax.grid(True, linestyle='--', alpha=0.7)
        
        # Ajouter une légende seulement si des courbes ont été tracées
        if legend_added:
            handles, labels = ax.get_legend_handles_labels()
            if handles:
                legend = ax.legend(handles, labels, fontsize=10, title="Simulations", loc="best")
        
        # Ajuster la mise en page
        self.gas_figure.tight_layout(pad=2.0)
        
        # Mettre à jour le canvas - utilisez le bon canvas
        self.gas_canvas.draw()
    
    def plot_liquid_graphs(self):
        """Afficher les graphiques de la phase liquide"""
        # Effacer le graphique actuel
        self.liquid_figure.clear()
        
        # Récupérer les options sélectionnées
        selected_options = [option for var, option in self.liquid_vars if var.get()]
        
        if not selected_options:
            messagebox.showinfo("Information", "Veuillez sélectionner au moins une courbe à afficher.")
            return
        
        # Créer un subplot unique
        ax = self.liquid_figure.add_subplot(111)
        
        # Palette de couleurs pour différentes simulations
        colors = plt.cm.tab10.colors
        
        # Tracer les courbes sélectionnées pour chaque simulation
        legend_added = False
        moisture_data = {}
        rh_data = {}
        
        # Collecter les données pour toutes les simulations
        for j, sim in enumerate(self.simulations):
            data = sim["data"]
            sim_name = sim["name"]
            color = colors[j % len(colors)]
            
            for option in selected_options:
                if option == "Humidité massique" and "MoistureFraction" in data:
                    ax.plot(data["Times"], data["MoistureFraction"], 
                           label=f"{sim_name} - Humidité massique", color=color, linestyle='-', linewidth=2)
                    legend_added = True
                    moisture_data[sim_name] = {
                        "times": data["Times"],
                        "values": data["MoistureFraction"],
                        "color": color
                    }
                    
                elif option == "Humidité relative" and "RelativeHumidity" in data:
                    ax.plot(data["Times"], data["RelativeHumidity"], 
                           label=f"{sim_name} - Humidité relative", color=color, linestyle='--', linewidth=2)
                    legend_added = True
                    rh_data[sim_name] = {
                        "times": data["Times"],
                        "values": data["RelativeHumidity"],
                        "color": color
                    }
        
        # Ajouter des annotations pour expliquer les zones d'humidité
        if moisture_data:
            # Prendre la première simulation pour l'analyse
            first_sim = list(moisture_data.keys())[0]
            times = moisture_data[first_sim]["times"]
            values = moisture_data[first_sim]["values"]
            
            # Définir les zones d'humidité de compostage importantes
            # Zone trop sèche: < 40%
            # Zone optimale: 40-65%
            # Zone trop humide: > 65%
            
            # Trouver les indices où l'humidité change de zone
            dry_zone = []
            optimal_zone = []
            wet_zone = []
            
            current_zone = None
            start_idx = 0
            
            for i, v in enumerate(values):
                if v < 40 and current_zone != "dry":
                    if current_zone is not None:
                        if current_zone == "optimal":
                            optimal_zone.append((start_idx, i-1))
                        elif current_zone == "wet":
                            wet_zone.append((start_idx, i-1))
                    current_zone = "dry"
                    start_idx = i
                elif 40 <= v <= 65 and current_zone != "optimal":
                    if current_zone is not None:
                        if current_zone == "dry":
                            dry_zone.append((start_idx, i-1))
                        elif current_zone == "wet":
                            wet_zone.append((start_idx, i-1))
                    current_zone = "optimal"
                    start_idx = i
                elif v > 65 and current_zone != "wet":
                    if current_zone is not None:
                        if current_zone == "dry":
                            dry_zone.append((start_idx, i-1))
                        elif current_zone == "optimal":
                            optimal_zone.append((start_idx, i-1))
                    current_zone = "wet"
                    start_idx = i
            
            # Ajouter la dernière zone
            if current_zone == "dry":
                dry_zone.append((start_idx, len(values)-1))
            elif current_zone == "optimal":
                optimal_zone.append((start_idx, len(values)-1))
            elif current_zone == "wet":
                wet_zone.append((start_idx, len(values)-1))
                
            # Colorer et annoter chaque zone
            for start, end in dry_zone:
                if end > start:
                    ax.axvspan(times[start], times[end], color='orange', alpha=0.1)
                    # Placer l'annotation au milieu de la zone
                    mid_idx = (start + end) // 2
                    ax.annotate('Zone trop sèche\n(<40%)\nActivité microbienne ralentie', 
                             xy=(times[mid_idx], values[mid_idx]),
                             xytext=(0, 20),
                             textcoords='offset points',
                             ha='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="orange", alpha=0.7))
            
            for start, end in optimal_zone:
                if end > start:
                    ax.axvspan(times[start], times[end], color='green', alpha=0.1)
                    # Placer l'annotation au milieu de la zone
                    mid_idx = (start + end) // 2
                    ax.annotate('Zone optimale\n(40-65%)\nBonne activité biologique', 
                             xy=(times[mid_idx], values[mid_idx]),
                             xytext=(0, -20),
                             textcoords='offset points',
                             ha='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="green", alpha=0.7))
            
            for start, end in wet_zone:
                if end > start:
                    ax.axvspan(times[start], times[end], color='blue', alpha=0.1)
                    # Placer l'annotation au milieu de la zone
                    mid_idx = (start + end) // 2
                    ax.annotate('Zone trop humide\n(>65%)\nRisque d\'anaérobie', 
                             xy=(times[mid_idx], values[mid_idx]),
                             xytext=(0, 20),
                             textcoords='offset points',
                             ha='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="blue", alpha=0.7))
            
            # Ajouter des lignes horizontales pour montrer les limites des zones
            ax.axhline(y=40, color='k', linestyle='--', alpha=0.5)
            ax.axhline(y=65, color='k', linestyle='--', alpha=0.5)
            
            # Ajouter du texte pour indiquer les limites des zones
            ax.text(times[-1], 40, ' 40%', va='center', ha='left', fontsize=9, bbox=dict(fc='white', ec='none', alpha=0.7))
            ax.text(times[-1], 65, ' 65%', va='center', ha='left', fontsize=9, bbox=dict(fc='white', ec='none', alpha=0.7))
            
            # Ajouter une légende pour les zones
            ax.legend(
                handles=[
                    mpatches.Patch(color='orange', alpha=0.1, label='Zone trop sèche (<40%)'),
                    mpatches.Patch(color='green', alpha=0.1, label='Zone optimale (40-65%)'),
                    mpatches.Patch(color='blue', alpha=0.1, label='Zone trop humide (>65%)')
                ],
                title="Zones d'humidité",
                loc="upper right",
                fontsize=9
            )
        
        # Si nous avons des données d'humidité relative, ajouter des zones et annotations
        elif rh_data:
            # Prendre la première simulation pour l'analyse
            first_sim = list(rh_data.keys())[0]
            times = rh_data[first_sim]["times"]
            values = rh_data[first_sim]["values"]
            
            # Zones d'humidité relative importantes
            # Zone sèche: < 70%
            # Zone normale: 70-90%
            # Zone saturée: > 90%
            
            # Dessiner les lignes horizontales pour les limites des zones
            ax.axhline(y=70, color='k', linestyle='--', alpha=0.5)
            ax.axhline(y=90, color='k', linestyle='--', alpha=0.5)
            
            # Ajouter du texte pour indiquer les limites des zones
            ax.text(times[-1], 70, ' 70%', va='center', ha='left', fontsize=9, bbox=dict(fc='white', ec='none', alpha=0.7))
            ax.text(times[-1], 90, ' 90%', va='center', ha='left', fontsize=9, bbox=dict(fc='white', ec='none', alpha=0.7))
            
            # Identifier les phases où l'humidité relative est dans différentes zones
            dry_zone = []
            normal_zone = []
            saturated_zone = []
            
            current_zone = None
            start_idx = 0
            
            for i, v in enumerate(values):
                if v < 70 and current_zone != "dry":
                    if current_zone is not None:
                        if current_zone == "normal":
                            normal_zone.append((start_idx, i-1))
                        elif current_zone == "saturated":
                            saturated_zone.append((start_idx, i-1))
                    current_zone = "dry"
                    start_idx = i
                elif 70 <= v <= 90 and current_zone != "normal":
                    if current_zone is not None:
                        if current_zone == "dry":
                            dry_zone.append((start_idx, i-1))
                        elif current_zone == "saturated":
                            saturated_zone.append((start_idx, i-1))
                    current_zone = "normal"
                    start_idx = i
                elif v > 90 and current_zone != "saturated":
                    if current_zone is not None:
                        if current_zone == "dry":
                            dry_zone.append((start_idx, i-1))
                        elif current_zone == "normal":
                            normal_zone.append((start_idx, i-1))
                    current_zone = "saturated"
                    start_idx = i
                    
            # Ajouter la dernière zone
            if current_zone == "dry":
                dry_zone.append((start_idx, len(values)-1))
            elif current_zone == "normal":
                normal_zone.append((start_idx, len(values)-1))
            elif current_zone == "saturated":
                saturated_zone.append((start_idx, len(values)-1))
                
            # Colorer et annoter chaque zone
            for start, end in dry_zone:
                if end > start:
                    ax.axvspan(times[start], times[end], color='orange', alpha=0.1)
                    mid_idx = (start + end) // 2
                    ax.annotate('Air sec (<70% HR)\nDessèchement possible', 
                             xy=(times[mid_idx], values[mid_idx]),
                             xytext=(0, 20),
                             textcoords='offset points',
                             ha='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="orange", alpha=0.7))
            
            for start, end in normal_zone:
                if end > start:
                    ax.axvspan(times[start], times[end], color='green', alpha=0.1)
                    mid_idx = (start + end) // 2
                    ax.annotate('Humidité normale (70-90% HR)', 
                             xy=(times[mid_idx], values[mid_idx]),
                             xytext=(0, -20),
                             textcoords='offset points',
                             ha='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="green", alpha=0.7))
            
            for start, end in saturated_zone:
                if end > start:
                    ax.axvspan(times[start], times[end], color='blue', alpha=0.1)
                    mid_idx = (start + end) // 2
                    ax.annotate('Air saturé (>90% HR)\nCondensat possible', 
                             xy=(times[mid_idx], values[mid_idx]),
                             xytext=(0, -20),
                             textcoords='offset points',
                             ha='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="blue", alpha=0.7))
        
        # Configurer le graphique
        ax.set_xlabel("Temps (h)", fontsize=12)
        ax.set_ylabel("Humidité (%)", fontsize=12)
        ax.set_title("Évolution des paramètres de la phase liquide", fontsize=14, fontweight='bold')
        ax.grid(True, linestyle='--', alpha=0.7)
        
        # Ajouter une légende pour les courbes
        if legend_added:
            handles, labels = ax.get_legend_handles_labels()
            if handles:
                legend = ax.legend(handles, labels, fontsize=10, title="Simulations", loc="upper left")
        
        # Ajuster la mise en page
        self.liquid_figure.tight_layout(pad=2.0)
        
        # Mettre à jour le canvas - Utiliser le bon canvas pour la phase liquide
        self.liquid_canvas.draw()
    
    def export_current_graph(self):
        """Exporter le graphique de l'onglet actif"""
        if not self.simulations:
            messagebox.showwarning("Avertissement", "Aucune simulation à exporter.")
            return
        
        # Déterminer l'onglet actif
        active_tab = self.phase_tabs.get()
        
        # Déterminer quelle figure exporter
        if active_tab == "Solide":
            figure = self.solid_figure
        elif active_tab == "Liquide":
            figure = self.liquid_figure
        elif active_tab == "Gazeuse":
            figure = self.gas_figure
        else:
            return
        
        # Demander où sauvegarder le fichier
        file_types = [
            ("Image PNG", "*.png"),
            ("Image JPEG", "*.jpg"),
            ("Document PDF", "*.pdf"),
            ("Données CSV", "*.csv"),
            ("Tous les fichiers", "*.*")
        ]
        
        file_path = ctk.filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=file_types,
            title=f"Exporter le graphique {active_tab}"
        )
        
        if not file_path:
            return  # L'utilisateur a annulé
        
        try:
            # Déterminer le format en fonction de l'extension
            extension = file_path.split(".")[-1].lower()
            
            if extension in ["png", "jpg", "jpeg", "pdf"]:
                # Exporter comme image
                figure.savefig(file_path, dpi=300, bbox_inches="tight")
                messagebox.showinfo("Succès", f"Graphique exporté avec succès vers {file_path}")
            elif extension == "csv":
                # Exporter les données sous-jacentes au format CSV
                import pandas as pd
                
                # Collecter les données de toutes les simulations
                all_data = {}
                for sim in self.simulations:
                    data = sim["data"]
                    sim_name = sim["name"]
                    
                    # Ajouter les données de temps
                    if "Times" in data:
                        all_data[f"Temps (h)"] = data["Times"]
                    
                    # Ajouter les données spécifiques selon l'onglet
                    if active_tab == "Solide":
                        if "Solids" in data:
                            all_data[f"{sim_name} - MS (kg)"] = data["Solids"]
                        if "Temperatures" in data:
                            all_data[f"{sim_name} - Température (°C)"] = data["Temperatures"]
                            
                        # Ajouter le rapport C/N
                        cn_ratios, success = self.calculate_cn_ratio(data)
                        if success:
                            all_data[f"{sim_name} - Rapport C/N"] = cn_ratios
                            
                    elif active_tab == "Liquide":
                        if "MoistureFraction" in data:
                            all_data[f"{sim_name} - Humidité (%)"] = data["MoistureFraction"]
                        if "RelativeHumidity" in data:
                            all_data[f"{sim_name} - HR (%)"] = data["RelativeHumidity"]
                    elif active_tab == "Gazeuse":
                        # Calculer les émissions de gaz
                        gas_data, success = self.calculate_gas_emissions(data)
                        if success:
                            if "CO2" in gas_data:
                                all_data[f"{sim_name} - CO₂ (kg)"] = gas_data["CO2"]
                            if "O2" in gas_data:
                                all_data[f"{sim_name} - O₂ (kg)"] = gas_data["O2"]
                            if "NH3" in gas_data:
                                all_data[f"{sim_name} - NH₃ (kg)"] = gas_data["NH3"]
                                
                        if "QExhaustgases" in data:
                            all_data[f"{sim_name} - Débit gaz (kg/h)"] = data["QExhaustgases"]
                        if "VExhaustgases" in data:
                            all_data[f"{sim_name} - Volume gaz (m³)"] = data["VExhaustgases"]
                
                # Créer et sauvegarder le DataFrame
                df = pd.DataFrame(all_data)
                df.to_csv(file_path, index=False)
                messagebox.showinfo("Succès", f"Données exportées avec succès vers {file_path}")
            else:
                messagebox.showwarning("Avertissement", f"Format de fichier {extension} non pris en charge.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'exportation: {str(e)}")
    
    def run_and_store_simulation(self):
        """Run simulation and store the results for comparison"""
        try:
            # Check if simulation parameters are set
            if not hasattr(self, 'simulation_params'):
                self.save_simulation_params()  # Use default values
                
            # Prepare data from the interface
            data = {
                "NS": self.NS,
                "Substrates": [sub["name"] for sub in self.substrates],
                "CCS": [sub["composition"] for sub in self.substrates],
                "FR": [sub["FR"] for sub in self.substrates],
                "FS": [sub["FS"] for sub in self.substrates],
                "FH": [100 - sub["FS"] for sub in self.substrates],  # FH = 100 - FS
                "FVS": [sub["FVS"] for sub in self.substrates],
                "FASH": [sub.get("FASH", 0) for sub in self.substrates],  # Default 0
                "FBVS": [sub["FBVS"] for sub in self.substrates],
                "FNBVS": [sub.get("FNBVS", 0) for sub in self.substrates],  # Default 0
                "FfBVS": [sub["FfBVS"] for sub in self.substrates],
                "FsBVS": [sub.get("FsBVS", 0) for sub in self.substrates],  # Default 0
                "T": [sub["T"] for sub in self.substrates],
                "fKT20": [sub["fKT20"] for sub in self.substrates],
                "sKT20": [sub["sKT20"] for sub in self.substrates],
                "Cp": [sub["Cp"] for sub in self.substrates],
                # Add simulation parameters
                "HRT": self.simulation_params["HRT"],
                "air_flow": self.simulation_params["air_flow"],
                "relative_humidity": self.simulation_params["relative_humidity"],
                "ambient_temp": self.simulation_params["ambient_temp"],
                "water_flow": self.simulation_params["water_flow"],
                "water_temp": self.simulation_params["water_temp"],
                "air_alternance": self.simulation_params["air_alternance"],
                "air_on_time": self.simulation_params["air_on_time"],
                "air_off_time": self.simulation_params["air_off_time"]
            }
            
            # Call process_data with prepared data
            sim_result = SimulationModel.run_simulation(data)
            
            # Store simulation with a name
            sim_name = self.sim_name_var.get()
            if not sim_name:
                sim_name = f"Simulation {len(self.simulations) + 1}"
            
            # Include simulation parameters in stored results
            sim_info = {
                "name": sim_name,
                "data": sim_result,
                "params": self.simulation_params.copy()
            }
            
            self.simulations.append(sim_info)
            self.simulation_names.append(sim_name)
            
            # Update simulation name for next run
            next_num = len(self.simulations) + 1
            self.sim_name_var.set(f"Simulation {next_num}")
            
            # Update the display of active simulations
            self.update_sim_display()
            
            # Update the data display with the new simulation results
            self.update_data_display()
            
            # Switch to the Data tab to show the updated results
            self.notebook.set("Données")
            
            messagebox.showinfo("Succès", f"Simulation '{sim_name}' terminée et sauvegardée. Les résultats sont affichés dans l'onglet Données.")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la simulation: {str(e)}")
    
    def update_sim_display(self):
        """Update the display showing active simulations"""
        self.sim_display.delete("1.0", "end")
        
        if not self.simulations:
            self.sim_display.insert("end", "Aucune simulation active. Lancez une simulation pour commencer.")
            return
        
        for i, sim in enumerate(self.simulations):
            alternance_info = "avec alternance" if sim["params"]["air_alternance"] else "sans alternance"
            air_flow = sim["params"]["air_flow"]
            hrt = sim["params"]["HRT"]
            
            self.sim_display.insert("end", f"{i+1}. {sim['name']} - Débit: {air_flow} m³/h, {alternance_info}, HRT: {hrt}h\n")
    
    def clear_simulations(self):
        """Clear all stored simulations"""
        if not self.simulations:
            messagebox.showinfo("Information", "Aucune simulation à effacer.")
            return
            
        if messagebox.askyesno("Confirmation", "Êtes-vous sûr de vouloir effacer toutes les simulations?"):
            self.simulations = []
            self.simulation_names = []
            self.sim_name_var.set("Simulation 1")
            self.update_sim_display()
            
            # Clear all plots
            self.solid_figure.clear()
            self.solid_canvas.draw()
            
            self.liquid_figure.clear()
            self.liquid_canvas.draw()
            
            self.gas_figure.clear()
            self.gas_canvas.draw()
            
            messagebox.showinfo("Succès", "Toutes les simulations ont été effacées.")
    
    def create_sensitivity_tab(self):
        """Create a tab for sensitivity analysis on air flow rate"""
        container = ctk.CTkFrame(self.sensitivity_frame)
        container.pack(expand=True, fill="both", padx=15, pady=15)
        
        # Title with background
        title_frame = ctk.CTkFrame(container, fg_color="#3a7ebf")
        title_frame.pack(fill="x", padx=10, pady=(0, 15))
        
        title_label = ctk.CTkLabel(
            title_frame, 
            text="Analyse de Sensibilité - Débit d'Air", 
            font=("Arial", 18, "bold"),
            text_color="white"
        )
        title_label.pack(pady=10)
        
        # Parameters frame
        params_frame = ctk.CTkFrame(container)
        params_frame.pack(fill="x", padx=10, pady=10)
        
        # Parameters layout with grid
        params_grid = ctk.CTkFrame(params_frame)
        params_grid.pack(fill="x", padx=10, pady=10)
        
        # Air flow range
        ctk.CTkLabel(params_grid, text="Intervalle de débit d'air:", font=("Arial", 12, "bold")).grid(row=0, column=0, padx=10, pady=10, sticky="w")
        
        range_frame = ctk.CTkFrame(params_grid)
        range_frame.grid(row=0, column=1, padx=10, pady=10, sticky="w")
        
        # Minimum value
        ctk.CTkLabel(range_frame, text="Min:").pack(side="left", padx=5)
        self.air_flow_min_var = ctk.StringVar(value="10.0")
        self.air_flow_min_entry = ctk.CTkEntry(range_frame, textvariable=self.air_flow_min_var, width=70)
        self.air_flow_min_entry.pack(side="left", padx=5)
        
        # Maximum value
        ctk.CTkLabel(range_frame, text="Max:").pack(side="left", padx=10)
        self.air_flow_max_var = ctk.StringVar(value="100.0")
        self.air_flow_max_entry = ctk.CTkEntry(range_frame, textvariable=self.air_flow_max_var, width=70)
        self.air_flow_max_entry.pack(side="left", padx=5)
        
        # Step value
        ctk.CTkLabel(range_frame, text="Pas:").pack(side="left", padx=10)
        self.air_flow_step_var = ctk.StringVar(value="10.0")
        self.air_flow_step_entry = ctk.CTkEntry(range_frame, textvariable=self.air_flow_step_var, width=70)
        self.air_flow_step_entry.pack(side="left", padx=5)
        
        ctk.CTkLabel(range_frame, text="m³/h").pack(side="left", padx=5)
        
        # Optimization criteria
        ctk.CTkLabel(params_grid, text="Critère d'optimisation:", font=("Arial", 12, "bold")).grid(row=1, column=0, padx=10, pady=10, sticky="w")
        
        criteria_frame = ctk.CTkFrame(params_grid)
        criteria_frame.grid(row=1, column=1, padx=10, pady=10, sticky="w")
        
        self.optimization_criteria = ctk.CTkComboBox(
            criteria_frame,
            values=["Température maximale", "Dégradation des solides", "Humidité finale", "Ratio dégradation/énergie", "Rapport C/N final"],
            width=200
        )
        self.optimization_criteria.pack(side="left", padx=5)
        
        # Run analysis button
        button_frame = ctk.CTkFrame(container)
        button_frame.pack(pady=20)
        
        run_analysis_btn = ctk.CTkButton(
            button_frame,
            text="Lancer l'analyse de sensibilité",
            command=self.run_sensitivity_analysis,
            width=250,
            height=40,
            font=("Arial", 14, "bold"),
            fg_color="#28a745",
            hover_color="#218838"
        )
        run_analysis_btn.pack(pady=5)
        
        # Results display area with scrollable canvas
        # Créer un cadre avec défilement
        results_container = ctk.CTkFrame(container)
        results_container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Créer un canvas scrollable
        self.sensitivity_canvas = ctk.CTkCanvas(results_container)
        scrollbar = ctk.CTkScrollbar(results_container, orientation="vertical", command=self.sensitivity_canvas.yview)
        scrollbar.pack(side="right", fill="y")
        
        self.sensitivity_canvas.configure(yscrollcommand=scrollbar.set)
        self.sensitivity_canvas.pack(side="left", fill="both", expand=True)
        
        # Cadre à l'intérieur du canvas pour contenir tous les éléments
        self.results_frame = ctk.CTkFrame(self.sensitivity_canvas)
        self.sensitivity_canvas.create_window((0, 0), window=self.results_frame, anchor="nw")
        
        # Configurer le défilement
        self.results_frame.bind("<Configure>", 
                           lambda e: self.sensitivity_canvas.configure(
                               scrollregion=self.sensitivity_canvas.bbox("all"),
                               width=e.width
                           ))
        
        # Activer le défilement avec la molette de souris
        self.sensitivity_canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        
        # Titre des résultats
        results_label = ctk.CTkLabel(
            self.results_frame,
            text="Résultats de l'analyse",
            font=("Arial", 14, "bold")
        )
        results_label.pack(anchor="w", padx=10, pady=10)
        
        # Zone pour afficher le graphique
        self.sensitivity_graph_frame = ctk.CTkFrame(self.results_frame)
        self.sensitivity_graph_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.sensitivity_figure = plt.Figure(figsize=(10, 10), dpi=100)  # Figure plus grande pour plus de détails
        self.sensitivity_canvas_plot = FigureCanvasTkAgg(self.sensitivity_figure, self.sensitivity_graph_frame)
        self.sensitivity_canvas_plot.get_tk_widget().pack(expand=True, fill="both")
        
        # Ajouter une barre d'outils de navigation pour le graphique
        toolbar_frame = ctk.CTkFrame(self.sensitivity_graph_frame)
        toolbar_frame.pack(side="bottom", fill="x")
        toolbar = NavigationToolbar2Tk(self.sensitivity_canvas_plot, toolbar_frame)
        toolbar.update()
        
        # Zone de texte pour les résultats
        self.sensitivity_results_text = ctk.CTkTextbox(self.results_frame, height=150)
        self.sensitivity_results_text.pack(fill="x", expand=False, padx=10, pady=10)
    
    def _on_mousewheel(self, event):
        """Gérer le défilement avec la molette de souris"""
        self.sensitivity_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    
    def calculate_cn_ratio(self, data, substrates=None):
        """Calcule le rapport C/N pour chaque pas de temps
        
        Args:
            data: Dictionnaire contenant les données de simulation
            substrates: Liste des substrats avec leurs compositions
            
        Returns:
            Liste des valeurs de rapport C/N pour chaque pas de temps
        """
        if not "Solids" in data or len(data["Solids"]) == 0:
            return [], False
            
        if substrates is None:
            substrates = self.substrates
            
        if not substrates:
            return [], False
            
        # Récupérer les données de dégradation
        times = data["Times"]
        solids = data["Solids"]
        
        # Initialiser le tableau pour les rapports C/N
        cn_ratios = []
        
        # Obtenir la composition initiale
        initial_c = 0
        initial_n = 0
        
        # Calculer la quantité initiale de C et N à partir des substrats
        for i, sub in enumerate(substrates):
            if i < len(substrates) and "composition" in sub:
                # La composition est [C, H, O, N]
                c_atoms = sub["composition"][0]  # Nombre d'atomes de carbone
                n_atoms = sub["composition"][3]  # Nombre d'atomes d'azote
                
                # Masses atomiques
                c_atomic_mass = 12.011  # g/mol
                n_atomic_mass = 14.007  # g/mol
                
                # Masse de C et N dans le substrat
                c_mass = c_atoms * c_atomic_mass
                n_mass = n_atoms * n_atomic_mass
                
                # Proportion massique de C et N dans le substrat
                total_mass = c_mass + n_mass
                c_fraction = c_mass / total_mass if total_mass > 0 else 0
                n_fraction = n_mass / total_mass if total_mass > 0 else 0
                
                # Masse initiale du substrat
                initial_solid = solids[0]
                
                # Contribution à la masse totale de C et N
                initial_c += initial_solid * c_fraction
                initial_n += initial_solid * n_fraction
        
        # Éviter la division par zéro
        if initial_n == 0:
            initial_n = 0.001
            
        # Calculer le rapport C/N initial
        initial_cn_ratio = initial_c / initial_n
        
        # Simuler la dégradation du rapport C/N
        for i in range(len(times)):
            # La dégradation affecte légèrement le rapport C/N
            # Dans un processus de compostage, le C est consommé plus rapidement que N
            degradation_factor = 1.0
            if i > 0:
                # Plus la dégradation des solides est importante, plus le rapport C/N diminue
                solid_degradation = (solids[0] - solids[i]) / solids[0] if solids[0] > 0 else 0
                # Le carbone se dégrade environ 30 fois plus vite que l'azote
                c_degradation = solid_degradation * 1.0  # Coefficient de dégradation du carbone
                n_degradation = solid_degradation * 0.03  # Coefficient de dégradation de l'azote
                
                # Calcul du nouveau rapport C/N
                remaining_c = initial_c * (1 - c_degradation)
                remaining_n = initial_n * (1 - n_degradation)
                
                # Éviter division par zéro
                if remaining_n <= 0:
                    remaining_n = 0.001
                    
                cn_ratio = remaining_c / remaining_n
            else:
                cn_ratio = initial_cn_ratio
                
            cn_ratios.append(cn_ratio)
        
        return cn_ratios, True

    def run_sensitivity_analysis(self):
        """Run sensitivity analysis with varying air flow rates"""
        try:
            # Vérifier si des substrats ont été définis
            if not self.substrates:
                messagebox.showerror("Erreur", "Aucun substrat défini. Veuillez d'abord configurer au moins un substrat.")
                return
            
            # Vérifier si les paramètres de simulation sont définis
            if not hasattr(self, 'simulation_params'):
                messagebox.showerror("Erreur", "Paramètres de simulation non définis. Veuillez d'abord enregistrer les paramètres de simulation.")
                return
            
            # Get the range values for air flow
            min_flow = float(self.air_flow_min_var.get())
            max_flow = float(self.air_flow_max_var.get())
            step_flow = float(self.air_flow_step_var.get())
            
            # Validate inputs
            if min_flow >= max_flow:
                messagebox.showerror("Erreur", "La valeur minimale doit être inférieure à la valeur maximale.")
                return
                
            if step_flow <= 0:
                messagebox.showerror("Erreur", "Le pas doit être supérieur à 0.")
                return
                
            # Generate the range of air flow values
            air_flow_values = [min_flow + i*step_flow for i in range(int((max_flow - min_flow) / step_flow) + 1) if min_flow + i*step_flow <= max_flow]
            
            # Get the selected optimization criteria
            criteria = self.optimization_criteria.get()
            
            # Show a progress dialog
            progress_window = ctk.CTkToplevel(self)
            progress_window.title("Analyse en cours")
            progress_window.geometry("400x150")
            progress_window.transient(self)
            progress_window.grab_set()
            
            prog_label = ctk.CTkLabel(progress_window, text="Exécution de l'analyse de sensibilité...", font=("Arial", 12))
            prog_label.pack(pady=15)
            
            progressbar = ctk.CTkProgressBar(progress_window, width=300)
            progressbar.pack(pady=10)
            progressbar.set(0)
            
            status_label = ctk.CTkLabel(progress_window, text="Initialisation...", font=("Arial", 10))
            status_label.pack(pady=5)
            
            progress_window.update()
            
            # Storage for results
            results = []
            criteria_values = []
            cn_ratio_values = []  # Pour stocker les rapports C/N finaux
            
            # Run simulations for each air flow value
            for i, flow in enumerate(air_flow_values):
                # Update progress
                progressbar.set((i+1)/len(air_flow_values))
                status_label.configure(text=f"Simulation {i+1}/{len(air_flow_values)} - Débit d'air: {flow} m³/h")
                progress_window.update()
                
                # Create a copy of simulation parameters with updated air flow
                sim_params = self.simulation_params.copy()
                sim_params["air_flow"] = flow
                
                # Prepare data from the interface
                data = {
                    "NS": self.NS,
                    "Substrates": [sub["name"] for sub in self.substrates],
                    "CCS": [sub["composition"] for sub in self.substrates],
                    "FR": [sub["FR"] for sub in self.substrates],
                    "FS": [sub["FS"] for sub in self.substrates],
                    "FH": [100 - sub["FS"] for sub in self.substrates],
                    "FVS": [sub["FVS"] for sub in self.substrates],
                    "FASH": [sub.get("FASH", 0) for sub in self.substrates],
                    "FBVS": [sub["FBVS"] for sub in self.substrates],
                    "FNBVS": [sub.get("FNBVS", 0) for sub in self.substrates],
                    "FfBVS": [sub["FfBVS"] for sub in self.substrates],
                    "FsBVS": [sub.get("FsBVS", 0) for sub in self.substrates],
                    "T": [sub["T"] for sub in self.substrates],
                    "fKT20": [sub["fKT20"] for sub in self.substrates],
                    "sKT20": [sub["sKT20"] for sub in self.substrates],
                    "Cp": [sub["Cp"] for sub in self.substrates],
                    # Utiliser les paramètres de simulation actuels
                    "HRT": sim_params["HRT"],
                    "air_flow": sim_params["air_flow"],  # Le débit d'air varie
                    "relative_humidity": sim_params["relative_humidity"],
                    "ambient_temp": sim_params["ambient_temp"],
                    "water_flow": sim_params["water_flow"],
                    "water_temp": sim_params["water_temp"],
                    "air_alternance": sim_params["air_alternance"],
                    "air_on_time": sim_params["air_on_time"],
                    "air_off_time": sim_params["air_off_time"]
                }
                
                # Run the simulation
                sim_result = SimulationModel.run_simulation(data)
                results.append(sim_result)
                
                # Calculer le rapport C/N final
                cn_ratios, success = self.calculate_cn_ratio(sim_result, self.substrates)
                if success and len(cn_ratios) > 0:
                    cn_ratio_values.append(cn_ratios[-1])  # Prendre la dernière valeur (finale)
                else:
                    cn_ratio_values.append(0)
                
                # Calculate the criteria value based on selected optimization criterion
                if criteria == "Température maximale":
                    criteria_value = max(sim_result["Temperatures"])
                elif criteria == "Dégradation des solides":
                    criteria_value = sim_result["Solids"][0] - sim_result["Solids"][-1]
                elif criteria == "Humidité finale":
                    criteria_value = sim_result["MoistureFraction"][-1]
                elif criteria == "Ratio dégradation/énergie":
                    solids_degraded = sim_result["Solids"][0] - sim_result["Solids"][-1]
                    energy_consumed = flow * sim_params["HRT"]  # Simple energy estimate
                    criteria_value = solids_degraded / energy_consumed if energy_consumed > 0 else 0
                elif criteria == "Rapport C/N final":
                    criteria_value = cn_ratios[-1] if success and len(cn_ratios) > 0 else 0
                else:
                    criteria_value = 0
                
                criteria_values.append(criteria_value)
            
            # Close the progress window
            progress_window.destroy()
            
            # Find the optimal air flow
            optimal_index = 0
            if criteria in ["Température maximale", "Dégradation des solides", "Ratio dégradation/énergie"]:
                # For these criteria, higher is better
                optimal_index = criteria_values.index(max(criteria_values))
            elif criteria == "Rapport C/N final":
                # Pour le rapport C/N, on cherche idéalement une valeur autour de 25-30
                ideal_cn = 25.0
                optimal_index = min(range(len(criteria_values)), key=lambda i: abs(criteria_values[i] - ideal_cn))
            else:
                # For humidity, closer to ideal (around 60%) is better
                ideal_humidity = 60.0
                optimal_index = min(range(len(criteria_values)), key=lambda i: abs(criteria_values[i] - ideal_humidity))
            
            optimal_flow = air_flow_values[optimal_index]
            
            # Plot the results
            self.sensitivity_figure.clear()
            
            # Définir une taille adaptée à la fenêtre
            self.sensitivity_figure.set_size_inches(10, 10)  # Dimensions plus grandes
            
            # Créer deux subplots dans la figure existante avec des proportions ajustées
            gs = self.sensitivity_figure.add_gridspec(2, 1, height_ratios=[3, 2], hspace=0.3)  # Plus d'espace entre les graphiques
            
            ax1 = self.sensitivity_figure.add_subplot(gs[0])  # Premier subplot avec plus d'espace
            ax2 = self.sensitivity_figure.add_subplot(gs[1], sharex=ax1)  # Deuxième subplot partageant l'axe x
            
            # Premier graphique - critère principal
            ax1.plot(air_flow_values, criteria_values, 'o-', linewidth=2.5, markersize=8, color='blue')
            ax1.set_ylabel(criteria, fontsize=12)
            ax1.set_title(f"Analyse de sensibilité - {criteria}", fontsize=14, fontweight='bold')
            ax1.grid(True, linestyle='--', alpha=0.7)
            ax1.tick_params(axis='both', which='major', labelsize=10)
            
            # Marquer le point optimal
            ax1.plot(optimal_flow, criteria_values[optimal_index], 'ro', markersize=12)
            ax1.annotate(f"Optimal: {optimal_flow:.1f} m³/h", 
                     (optimal_flow, criteria_values[optimal_index]),
                     xytext=(15, 15),
                     textcoords='offset points',
                     fontsize=11,
                     bbox=dict(boxstyle="round,pad=0.3", fc="yellow", ec="black", alpha=0.8),
                     arrowprops=dict(arrowstyle="->", connectionstyle="arc3,rad=.5"))
            
            # Ajouter des zones explicatives pour le graphique principal
            if criteria in ["Dégradation des solides", "Température maximale", "Ratio dégradation/énergie"]:
                # Zone de croissance
                ax1.axvspan(min(air_flow_values), optimal_flow, color='blue', alpha=0.1)
                ax1.annotate('Zone efficace\n(rendement croissant)', 
                         ((min(air_flow_values) + optimal_flow)/2, max(criteria_values)*0.8),
                         ha='center', va='center',
                         fontsize=10,
                         bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="blue", alpha=0.8))
                
                # Zone de décroissance (après l'optimal)
                ax1.axvspan(optimal_flow, max(air_flow_values), color='red', alpha=0.1)
                ax1.annotate('Zone inefficace\n(rendement décroissant)', 
                         ((optimal_flow + max(air_flow_values))/2, max(criteria_values)*0.7),
                         ha='center', va='center',
                         fontsize=10,
                         bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="red", alpha=0.8))
            elif criteria == "Humidité finale":
                # Pour l'humidité finale, on cherche souvent une valeur optimale autour de 60%
                ideal_humidity = 60.0
                
                # Zone trop sèche
                if min(criteria_values) < ideal_humidity:
                    ax1.axvspan(min(air_flow_values), optimal_flow, color='orange', alpha=0.1)
                    ax1.annotate('Zone trop sèche\n(<60%)', 
                             ((min(air_flow_values) + optimal_flow)/2, min(criteria_values)*1.1),
                             ha='center', va='center',
                             fontsize=10,
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="orange", alpha=0.8))
                
                # Zone trop humide
                if max(criteria_values) > ideal_humidity:
                    ax1.axvspan(optimal_flow, max(air_flow_values), color='blue', alpha=0.1)
                    ax1.annotate('Zone trop humide\n(>60%)', 
                             ((optimal_flow + max(air_flow_values))/2, max(criteria_values)*0.9),
                             ha='center', va='center',
                             fontsize=10,
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="blue", alpha=0.8))
            
            # Deuxième graphique - rapport C/N
            ax2.plot(air_flow_values, cn_ratio_values, 'o-', linewidth=2.5, markersize=8, color='green')
            ax2.set_xlabel("Débit d'air (m³/h)", fontsize=12)
            ax2.set_ylabel("Rapport C/N", fontsize=12)
            ax2.grid(True, linestyle='--', alpha=0.7)
            ax2.tick_params(axis='both', which='major', labelsize=10)
            
            # Ajouter un titre au deuxième graphique
            ax2.set_title("Impact du débit d'air sur le rapport C/N", fontsize=12)
            
            # Marquer la zone optimale pour le rapport C/N (20-30)
            ax2.axhspan(20, 30, color='green', alpha=0.2, label='Zone optimale C/N')
            ax2.annotate('Zone optimale C/N (20-30)', 
                     (air_flow_values[len(air_flow_values)//2], 25),
                     ha='center', va='center',
                     fontsize=10,
                     bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="green", alpha=0.8))
            
            # Zone sub-optimale pour le rapport C/N (<20)
            if min(cn_ratio_values) < 20:
                ax2.axhspan(0, 20, color='orange', alpha=0.1, label='Zone C/N faible')
                ax2.annotate('C/N trop faible (<20)\nDécomposition rapide, risque de perte d\'azote', 
                         (air_flow_values[len(air_flow_values)//4], 10),
                         ha='center', va='center',
                         fontsize=9,
                         bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="orange", alpha=0.8))
            
            # Zone sub-optimale pour le rapport C/N (>30)
            if max(cn_ratio_values) > 30:
                ax2.axhspan(30, max(cn_ratio_values) + 5, color='red', alpha=0.1, label='Zone C/N élevé')
                ax2.annotate('C/N trop élevé (>30)\nDécomposition lente', 
                         (air_flow_values[3*len(air_flow_values)//4], min(max(cn_ratio_values), 40)),
                         ha='center', va='center',
                         fontsize=9,
                         bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="red", alpha=0.8))
            
            # Ajouter une légende personnalisée pour expliquer la signification des zones
            legend_elements = []
            
            if criteria in ["Dégradation des solides", "Température maximale", "Ratio dégradation/énergie"]:
                legend_elements.append(mpatches.Patch(color='blue', alpha=0.2, label='Zone efficace'))
                legend_elements.append(mpatches.Patch(color='red', alpha=0.2, label='Zone inefficace'))
            elif criteria == "Humidité finale":
                legend_elements.append(mpatches.Patch(color='orange', alpha=0.2, label='Zone trop sèche'))
                legend_elements.append(mpatches.Patch(color='blue', alpha=0.2, label='Zone trop humide'))
            
            legend_elements.append(mpatches.Patch(color='green', alpha=0.2, label='Zone C/N optimale (20-30)'))
            
            if min(cn_ratio_values) < 20:
                legend_elements.append(mpatches.Patch(color='orange', alpha=0.2, label='Zone C/N faible (<20)'))
            if max(cn_ratio_values) > 30:
                legend_elements.append(mpatches.Patch(color='red', alpha=0.2, label='Zone C/N élevé (>30)'))
            
            # Ajouter la légende
            self.sensitivity_figure.legend(
                handles=legend_elements,
                loc='upper center', 
                bbox_to_anchor=(0.5, 0.98),
                ncol=min(4, len(legend_elements)),
                fontsize=9,
                fancybox=True, 
                shadow=True
            )
            
            # Marquer également le point optimal sur le graphique du rapport C/N
            ax2.plot(optimal_flow, cn_ratio_values[optimal_index], 'ro', markersize=10)
            ax2.annotate(f"C/N: {cn_ratio_values[optimal_index]:.1f}", 
                     (optimal_flow, cn_ratio_values[optimal_index]),
                     xytext=(15, 10),
                     textcoords='offset points',
                     fontsize=9,
                     bbox=dict(boxstyle="round,pad=0.2", fc="yellow", ec="black", alpha=0.7),
                     arrowprops=dict(arrowstyle="->", connectionstyle="arc3,rad=.5"))
                        
            # Ajuster la mise en page avec marges réduites pour assurer la visibilité complète
            plt.tight_layout(rect=[0.03, 0.03, 0.97, 0.92], pad=0.4, h_pad=0.8, w_pad=0.5)
            
            # Mettre à jour le canvas existant
            self.sensitivity_canvas_plot.draw()
            
            # Display summary in text area
            self.sensitivity_results_text.delete("1.0", "end")
            self.sensitivity_results_text.insert("end", f"=== RÉSULTATS DE L'ANALYSE DE SENSIBILITÉ ===\n\n")
            self.sensitivity_results_text.insert("end", f"Critère d'optimisation: {criteria}\n")
            self.sensitivity_results_text.insert("end", f"Débit d'air optimal: {optimal_flow:.1f} m³/h\n")
            
            if criteria == "Température maximale":
                self.sensitivity_results_text.insert("end", f"Température maximale atteinte: {criteria_values[optimal_index]:.2f} °C\n")
            elif criteria == "Dégradation des solides":
                self.sensitivity_results_text.insert("end", f"Dégradation totale des solides: {criteria_values[optimal_index]:.2f} kg\n")
            elif criteria == "Humidité finale":
                self.sensitivity_results_text.insert("end", f"Humidité finale: {criteria_values[optimal_index]:.2f}%\n")
            elif criteria == "Ratio dégradation/énergie":
                self.sensitivity_results_text.insert("end", f"Ratio dégradation/énergie: {criteria_values[optimal_index]:.4f} kg/m³\n")
            elif criteria == "Rapport C/N final":
                self.sensitivity_results_text.insert("end", f"Rapport C/N final: {criteria_values[optimal_index]:.2f}\n")
            
            # Ajouter toujours l'information sur le rapport C/N final
            if criteria != "Rapport C/N final":
                self.sensitivity_results_text.insert("end", f"Rapport C/N final: {cn_ratio_values[optimal_index]:.2f}\n")
            
            # Ajouter des informations sur les paramètres utilisés pour la simulation
            self.sensitivity_results_text.insert("end", "\n=== PARAMÈTRES UTILISÉS ===\n")
            self.sensitivity_results_text.insert("end", f"HRT: {sim_params['HRT']} heures\n")
            self.sensitivity_results_text.insert("end", f"Humidité relative: {sim_params['relative_humidity']}%\n")
            self.sensitivity_results_text.insert("end", f"Température ambiante: {sim_params['ambient_temp']}°C\n")
            self.sensitivity_results_text.insert("end", f"Débit d'eau: {sim_params['water_flow']} kg/h\n")
            
            # Add the optimal simulation parameters to the app
            optimal_params = self.simulation_params.copy()
            optimal_params["air_flow"] = optimal_flow
            
            # Ask if user wants to set this as the current flow rate
            if messagebox.askyesno("Optimisation", f"Voulez-vous définir le débit d'air optimal ({optimal_flow:.1f} m³/h) comme valeur par défaut pour les simulations?"):
                self.simulation_params["air_flow"] = optimal_flow
                self.air_flow_var.set(str(optimal_flow))
                messagebox.showinfo("Mise à jour", "Le débit d'air a été mis à jour avec la valeur optimale.")
                
            # Ajouter une légende personnalisée pour expliquer la signification des zones
            blue_patch = mpatches.Patch(color='blue', alpha=0.1, label='Zone de croissance')
            red_patch = mpatches.Patch(color='red', alpha=0.1, label='Zone de décroissance')
            green_patch = mpatches.Patch(color='green', alpha=0.2, label='Zone C/N optimale (20-30)')
            orange_patch = mpatches.Patch(color='orange', alpha=0.1, label='Zone C/N faible (<20)')
            
            # Ajouter une légende commune dans le haut du graphique
            self.sensitivity_figure.legend(
                handles=[blue_patch, red_patch, green_patch, orange_patch],
                loc='upper center', 
                bbox_to_anchor=(0.5, 0.97),
                ncol=4,
                fontsize=8,
                fancybox=True, 
                shadow=True
            )
            
            # Ajuster la mise en page avec la légende
            plt.tight_layout(rect=[0, 0, 1, 0.92], pad=0.4, h_pad=0.5, w_pad=0.5)  # Réserver de l'espace en haut pour la légende
            
            # Mettre à jour le canvas existant
            self.sensitivity_canvas.draw()
            
                        
            # Ajuster la mise en page avec marges réduites pour assurer la visibilité complète
            plt.tight_layout(rect=[0, 0, 1, 0.95], pad=0.3, h_pad=0.5, w_pad=0.5)
            
            # Définir une taille encore plus compacte pour la fenêtre
            self.sensitivity_figure.set_size_inches(7, 5)  # Dimensions réduites
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'analyse de sensibilité: {str(e)}")
    
    def calculate_gas_emissions(self, data):
        """Calcule les émissions de gaz (CO2, O2, NH3) basées sur la dégradation de la matière organique"""
        # Vérifier si les données nécessaires sont disponibles
        if not all(key in data for key in ["Times", "Solids"]):
            return {}, False
            
        # Récupérer les données de dégradation
        times = data["Times"]
        initial_solids = data["Solids"][0]
        final_solids = data["Solids"][-1]
        solids = data["Solids"]
        
        # Initialiser les tableaux pour stocker les valeurs calculées
        co2_values = []
        o2_values = []
        nh3_values = []
        
        # Facteurs stœchiométriques approximatifs pour la dégradation aérobie
        # Basés sur la composition typique de la matière organique C₂₇H₃₈O₁₆N
        CO2_FACTOR = 1.5  # kg CO2 / kg matière organique dégradée
        O2_FACTOR = 1.2   # kg O2 / kg matière organique dégradée
        NH3_FACTOR = 0.05 # kg NH3 / kg matière organique dégradée
        
        # Calcul cumulatif des émissions
        cumulative_co2 = 0
        cumulative_o2 = 0
        cumulative_nh3 = 0
        
        for i in range(len(times)):
            if i > 0:
                # Calcul de la matière dégradée entre ce pas de temps et le précédent
                degraded = solids[i-1] - solids[i]
                if degraded > 0:
                    # Conversion de la matière dégradée en gaz émis
                    co2_produced = degraded * CO2_FACTOR
                    o2_consumed = degraded * O2_FACTOR
                    nh3_produced = degraded * NH3_FACTOR
                    
                    # Accumulation
                    cumulative_co2 += co2_produced
                    cumulative_o2 += o2_consumed
                    cumulative_nh3 += nh3_produced
            
            # Stocker les valeurs cumulatives
            co2_values.append(cumulative_co2)
            o2_values.append(cumulative_o2)
            nh3_values.append(cumulative_nh3)
        
        # Retourner les résultats calculés
        gas_data = {
            "CO2": co2_values,
            "O2": o2_values,
            "NH3": nh3_values
        }
        
        return gas_data, True

    def plot_solid_graphs(self):
        """Afficher les graphiques de la phase solide"""
        try:
            if not self.simulations:
                messagebox.showwarning("Avertissement", "Aucune simulation à afficher. Veuillez d'abord lancer une simulation.")
                return
            
            # Effacer le graphique actuel
            self.solid_figure.clear()
            
            # Récupérer les options sélectionnées
            selected_options = [option for var, option in self.solid_vars if var.get()]
            
            if not selected_options:
                messagebox.showinfo("Information", "Veuillez sélectionner au moins une courbe à afficher.")
                return
            
            # Créer un subplot unique
            ax = self.solid_figure.add_subplot(111)
            
            # Palette de couleurs pour différentes simulations
            colors = plt.cm.tab10.colors
            
            # Tracer les courbes sélectionnées pour chaque simulation
            legend_added = False
            temperatures_data = {}
            solids_data = {}
            cn_data = {}
            
            # Collecter d'abord toutes les données pour pouvoir analyser les phases
            for j, sim in enumerate(self.simulations):
                data = sim["data"]
                sim_name = sim["name"]
                color = colors[j % len(colors)]
                
                for option in selected_options:
                    if option == "Matière sèche (MS)" and "Solids" in data:
                        ax.plot(data["Times"], data["Solids"], 
                               label=f"{sim_name} - MS", color=color, linestyle='-', linewidth=2)
                        legend_added = True
                        solids_data[sim_name] = {
                            "times": data["Times"],
                            "values": data["Solids"],
                            "color": color
                        }
                        
                    elif option == "Température" and "Temperatures" in data:
                        ax.plot(data["Times"], data["Temperatures"], 
                               label=f"{sim_name} - Température", color=color, linestyle='--', linewidth=2)
                        legend_added = True
                        temperatures_data[sim_name] = {
                            "times": data["Times"],
                            "values": data["Temperatures"],
                            "color": color
                        }
                        
                    elif option == "Rapport C/N":
                        # Calculer le rapport C/N
                        cn_ratios, success = self.calculate_cn_ratio(data)
                        if success:
                            ax.plot(data["Times"], cn_ratios, 
                                   label=f"{sim_name} - C/N", color=color, linestyle='-.', linewidth=2)
                            legend_added = True
                            cn_data[sim_name] = {
                                "times": data["Times"],
                                "values": cn_ratios,
                                "color": color
                            }
            
            # Ajouter des annotations pour expliquer les différentes phases du processus
            # Nous allons analyser la courbe de température pour identifier les phases
            if temperatures_data:
                # Prendre la première simulation pour l'analyse des phases
                first_sim = list(temperatures_data.keys())[0]
                temp_times = temperatures_data[first_sim]["times"]
                temp_values = temperatures_data[first_sim]["values"]
                
                # Identifier les phases caractéristiques du compostage
                # Phase 1: Phase mésophile (T < 45°C) - début du processus
                # Phase 2: Phase thermophile (T > 45°C) - activité biologique intense
                # Phase 3: Phase de refroidissement - diminution de l'activité
                # Phase 4: Phase de maturation - stabilisation
                
                max_temp_idx = temp_values.index(max(temp_values))
                max_temp_time = temp_times[max_temp_idx]
                
                # Trouver l'indice où la température dépasse 45°C
                try:
                    thermo_start_idx = next(i for i, t in enumerate(temp_values) if t >= 45)
                    thermo_start_time = temp_times[thermo_start_idx]
                except (StopIteration, IndexError):
                    thermo_start_idx = max_temp_idx // 2
                    thermo_start_time = temp_times[thermo_start_idx]
                
                # Trouver l'indice où la température redescend sous 45°C après la phase thermophile
                cooling_start_idx = max_temp_idx
                for i in range(max_temp_idx, len(temp_values)):
                    if temp_values[i] < 45:
                        cooling_start_idx = i
                        break
                cooling_start_time = temp_times[cooling_start_idx]
                
                # Maturation: dernier tiers après la phase de refroidissement
                maturation_start_idx = (cooling_start_idx + len(temp_values)) // 2
                maturation_start_time = temp_times[maturation_start_idx]
                
                # Délimiter et annoter les phases
                # Phase mésophile initiale
                if thermo_start_idx > 0:
                    ax.axvspan(temp_times[0], thermo_start_time, color='blue', alpha=0.1)
                    ax.annotate('Phase mésophile\n(Démarrage)', 
                             xy=(temp_times[0] + (thermo_start_time - temp_times[0])/2, min(temp_values) + 10),
                             ha='center', va='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="blue", alpha=0.7))
                
                # Phase thermophile
                if thermo_start_idx < cooling_start_idx:
                    ax.axvspan(thermo_start_time, cooling_start_time, color='red', alpha=0.1)
                    ax.annotate('Phase thermophile\n(Haute activité biologique)', 
                             xy=(thermo_start_time + (cooling_start_time - thermo_start_time)/2, max(temp_values) - 5),
                             ha='center', va='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="red", alpha=0.7))
                
                # Phase de refroidissement
                if cooling_start_idx < maturation_start_idx:
                    ax.axvspan(cooling_start_time, maturation_start_time, color='green', alpha=0.1)
                    ax.annotate('Phase de refroidissement', 
                             xy=(cooling_start_time + (maturation_start_time - cooling_start_time)/2, 
                                 (min(temp_values) + max(temp_values))/2),
                             ha='center', va='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="green", alpha=0.7))
                
                # Phase de maturation
                if maturation_start_idx < len(temp_times) - 1:
                    ax.axvspan(maturation_start_time, temp_times[-1], color='purple', alpha=0.1)
                    ax.annotate('Phase de maturation\n(Stabilisation)', 
                             xy=(maturation_start_time + (temp_times[-1] - maturation_start_time)/2, 
                                 min(temp_values) + 5),
                             ha='center', va='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="purple", alpha=0.7))
                    
                # Ajouter une légende pour les phases
                ax.legend(
                    handles=[
                        mpatches.Patch(color='blue', alpha=0.1, label='Phase mésophile'),
                        mpatches.Patch(color='red', alpha=0.1, label='Phase thermophile'),
                        mpatches.Patch(color='green', alpha=0.1, label='Phase de refroidissement'),
                        mpatches.Patch(color='purple', alpha=0.1, label='Phase de maturation')
                    ],
                    title="Phases du processus",
                    loc="upper right",
                    fontsize=9
                )
            
            # Si aucune courbe de température mais des données de matière sèche
            elif solids_data:
                # Analyser la dégradation des solides
                first_sim = list(solids_data.keys())[0]
                solid_times = solids_data[first_sim]["times"]
                solid_values = solids_data[first_sim]["values"]
                
                # Calculer le taux de dégradation
                degradation_rates = [0]
                for i in range(1, len(solid_values)):
                    rate = (solid_values[i-1] - solid_values[i]) / (solid_times[i] - solid_times[i-1])
                    degradation_rates.append(rate)
                
                # Lisser les taux pour l'analyse
                from scipy.signal import savgol_filter
                try:
                    smooth_rates = savgol_filter(degradation_rates, min(11, len(degradation_rates) - 1 if len(degradation_rates) % 2 == 0 else len(degradation_rates)), 3)
                except:
                    smooth_rates = degradation_rates
                
                # Identifier les phases
                max_rate_idx = list(smooth_rates).index(max(smooth_rates[1:]))
                max_rate_time = solid_times[max_rate_idx]
                
                # Phase initiale: avant le taux de dégradation maximal
                initial_end_idx = max(1, max_rate_idx)
                initial_end_time = solid_times[initial_end_idx]
                
                # Phase de dégradation active: autour du taux maximal
                active_end_idx = min(len(solid_times)-1, max_rate_idx + len(solid_times)//3)
                active_end_time = solid_times[active_end_idx]
                
                # Phase de stabilisation: dernière partie
                if active_end_idx < len(solid_times) - 1:
                    ax.axvspan(solid_times[0], initial_end_time, color='blue', alpha=0.1)
                    ax.annotate('Phase d\'initialisation', 
                             xy=(solid_times[0] + (initial_end_time - solid_times[0])/2, solid_values[0] - 5),
                             ha='center', va='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="blue", alpha=0.7))
                    
                    ax.axvspan(initial_end_time, active_end_time, color='red', alpha=0.1)
                    ax.annotate('Dégradation active', 
                             xy=(initial_end_time + (active_end_time - initial_end_time)/2, 
                                 (solid_values[initial_end_idx] + solid_values[active_end_idx])/2),
                             ha='center', va='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="red", alpha=0.7))
                    
                    ax.axvspan(active_end_time, solid_times[-1], color='green', alpha=0.1)
                    ax.annotate('Stabilisation', 
                             xy=(active_end_time + (solid_times[-1] - active_end_time)/2, 
                                 solid_values[-1] + 5),
                             ha='center', va='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="green", alpha=0.7))
            
            # Configurer le graphique
            ax.set_xlabel("Temps (h)", fontsize=12)
            
            # Définir l'unité appropriée en fonction des courbes sélectionnées
            if len(selected_options) == 1:
                if "Matière sèche (MS)" in selected_options:
                    ax.set_ylabel("Quantité (kg)", fontsize=12)
                elif "Température" in selected_options:
                    ax.set_ylabel("Température (°C)", fontsize=12)
                elif "Rapport C/N" in selected_options:
                    ax.set_ylabel("Rapport C/N", fontsize=12)
            else:
                # Si plusieurs courbes sont sélectionnées avec des unités différentes
                contains_ms = "Matière sèche (MS)" in selected_options
                contains_temp = "Température" in selected_options
                contains_cn = "Rapport C/N" in selected_options
                
                if contains_ms and not contains_temp and not contains_cn:
                    ax.set_ylabel("Quantité (kg)", fontsize=12)
                elif contains_temp and not contains_ms and not contains_cn:
                    ax.set_ylabel("Température (°C)", fontsize=12)
                elif contains_cn and not contains_ms and not contains_temp:
                    ax.set_ylabel("Rapport C/N", fontsize=12)
                else:
                    # Plusieurs types de données avec différentes unités
                    ax.set_ylabel("Valeur (voir légende pour unités)", fontsize=12)
            
            ax.set_title("Évolution des paramètres de la phase solide", fontsize=14, fontweight='bold')
            ax.grid(True, linestyle='--', alpha=0.7)
            
            # Ajouter une légende seulement si des courbes ont été tracées
            if legend_added:
                handles, labels = ax.get_legend_handles_labels()
                if handles:
                    legend = ax.legend(handles, labels, fontsize=10, title="Simulations", loc="best")
            
            # Ajuster la mise en page
            self.solid_figure.tight_layout(pad=2.0)
            
            # Mettre à jour le canvas - Utiliser le bon canvas pour la phase solide
            self.solid_canvas.draw()
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'affichage du graphique: {str(e)}")
            import traceback
            traceback.print_exc()

    def plot_liquid_graphs(self):
        """Afficher les graphiques de la phase liquide"""
        # Effacer le graphique actuel
        self.liquid_figure.clear()
        
        # Récupérer les options sélectionnées
        selected_options = [option for var, option in self.liquid_vars if var.get()]
        
        if not selected_options:
            messagebox.showinfo("Information", "Veuillez sélectionner au moins une courbe à afficher.")
            return
        
        # Créer un subplot unique
        ax = self.liquid_figure.add_subplot(111)
        
        # Palette de couleurs pour différentes simulations
        colors = plt.cm.tab10.colors
        
        # Tracer les courbes sélectionnées pour chaque simulation
        legend_added = False
        moisture_data = {}
        rh_data = {}
        
        # Collecter les données pour toutes les simulations
        for j, sim in enumerate(self.simulations):
            data = sim["data"]
            sim_name = sim["name"]
            color = colors[j % len(colors)]
            
            for option in selected_options:
                if option == "Humidité massique" and "MoistureFraction" in data:
                    ax.plot(data["Times"], data["MoistureFraction"], 
                           label=f"{sim_name} - Humidité massique", color=color, linestyle='-', linewidth=2)
                    legend_added = True
                    moisture_data[sim_name] = {
                        "times": data["Times"],
                        "values": data["MoistureFraction"],
                        "color": color
                    }
                    
                elif option == "Humidité relative" and "RelativeHumidity" in data:
                    ax.plot(data["Times"], data["RelativeHumidity"], 
                           label=f"{sim_name} - Humidité relative", color=color, linestyle='--', linewidth=2)
                    legend_added = True
                    rh_data[sim_name] = {
                        "times": data["Times"],
                        "values": data["RelativeHumidity"],
                        "color": color
                    }
        
        # Ajouter des annotations pour expliquer les zones d'humidité
        if moisture_data:
            # Prendre la première simulation pour l'analyse
            first_sim = list(moisture_data.keys())[0]
            times = moisture_data[first_sim]["times"]
            values = moisture_data[first_sim]["values"]
            
            # Définir les zones d'humidité de compostage importantes
            # Zone trop sèche: < 40%
            # Zone optimale: 40-65%
            # Zone trop humide: > 65%
            
            # Trouver les indices où l'humidité change de zone
            dry_zone = []
            optimal_zone = []
            wet_zone = []
            
            current_zone = None
            start_idx = 0
            
            for i, v in enumerate(values):
                if v < 40 and current_zone != "dry":
                    if current_zone is not None:
                        if current_zone == "optimal":
                            optimal_zone.append((start_idx, i-1))
                        elif current_zone == "wet":
                            wet_zone.append((start_idx, i-1))
                    current_zone = "dry"
                    start_idx = i
                elif 40 <= v <= 65 and current_zone != "optimal":
                    if current_zone is not None:
                        if current_zone == "dry":
                            dry_zone.append((start_idx, i-1))
                        elif current_zone == "wet":
                            wet_zone.append((start_idx, i-1))
                    current_zone = "optimal"
                    start_idx = i
                elif v > 65 and current_zone != "wet":
                    if current_zone is not None:
                        if current_zone == "dry":
                            dry_zone.append((start_idx, i-1))
                        elif current_zone == "optimal":
                            optimal_zone.append((start_idx, i-1))
                    current_zone = "wet"
                    start_idx = i
            
            # Ajouter la dernière zone
            if current_zone == "dry":
                dry_zone.append((start_idx, len(values)-1))
            elif current_zone == "optimal":
                optimal_zone.append((start_idx, len(values)-1))
            elif current_zone == "wet":
                wet_zone.append((start_idx, len(values)-1))
                
            # Colorer et annoter chaque zone
            for start, end in dry_zone:
                if end > start:
                    ax.axvspan(times[start], times[end], color='orange', alpha=0.1)
                    # Placer l'annotation au milieu de la zone
                    mid_idx = (start + end) // 2
                    ax.annotate('Zone trop sèche\n(<40%)\nActivité microbienne ralentie', 
                             xy=(times[mid_idx], values[mid_idx]),
                             xytext=(0, 20),
                             textcoords='offset points',
                             ha='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="orange", alpha=0.7))
            
            for start, end in optimal_zone:
                if end > start:
                    ax.axvspan(times[start], times[end], color='green', alpha=0.1)
                    # Placer l'annotation au milieu de la zone
                    mid_idx = (start + end) // 2
                    ax.annotate('Zone optimale\n(40-65%)\nBonne activité biologique', 
                             xy=(times[mid_idx], values[mid_idx]),
                             xytext=(0, -20),
                             textcoords='offset points',
                             ha='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="green", alpha=0.7))
            
            for start, end in wet_zone:
                if end > start:
                    ax.axvspan(times[start], times[end], color='blue', alpha=0.1)
                    # Placer l'annotation au milieu de la zone
                    mid_idx = (start + end) // 2
                    ax.annotate('Zone trop humide\n(>65%)\nRisque d\'anaérobie', 
                             xy=(times[mid_idx], values[mid_idx]),
                             xytext=(0, 20),
                             textcoords='offset points',
                             ha='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="blue", alpha=0.7))
            
            # Ajouter des lignes horizontales pour montrer les limites des zones
            ax.axhline(y=40, color='k', linestyle='--', alpha=0.5)
            ax.axhline(y=65, color='k', linestyle='--', alpha=0.5)
            
            # Ajouter du texte pour indiquer les limites des zones
            ax.text(times[-1], 40, ' 40%', va='center', ha='left', fontsize=9, bbox=dict(fc='white', ec='none', alpha=0.7))
            ax.text(times[-1], 65, ' 65%', va='center', ha='left', fontsize=9, bbox=dict(fc='white', ec='none', alpha=0.7))
            
            # Ajouter une légende pour les zones
            ax.legend(
                handles=[
                    mpatches.Patch(color='orange', alpha=0.1, label='Zone trop sèche (<40%)'),
                    mpatches.Patch(color='green', alpha=0.1, label='Zone optimale (40-65%)'),
                    mpatches.Patch(color='blue', alpha=0.1, label='Zone trop humide (>65%)')
                ],
                title="Zones d'humidité",
                loc="upper right",
                fontsize=9
            )
        
        # Si nous avons des données d'humidité relative, ajouter des zones et annotations
        elif rh_data:
            # Prendre la première simulation pour l'analyse
            first_sim = list(rh_data.keys())[0]
            times = rh_data[first_sim]["times"]
            values = rh_data[first_sim]["values"]
            
            # Zones d'humidité relative importantes
            # Zone sèche: < 70%
            # Zone normale: 70-90%
            # Zone saturée: > 90%
            
            # Dessiner les lignes horizontales pour les limites des zones
            ax.axhline(y=70, color='k', linestyle='--', alpha=0.5)
            ax.axhline(y=90, color='k', linestyle='--', alpha=0.5)
            
            # Ajouter du texte pour indiquer les limites des zones
            ax.text(times[-1], 70, ' 70%', va='center', ha='left', fontsize=9, bbox=dict(fc='white', ec='none', alpha=0.7))
            ax.text(times[-1], 90, ' 90%', va='center', ha='left', fontsize=9, bbox=dict(fc='white', ec='none', alpha=0.7))
            
            # Identifier les phases où l'humidité relative est dans différentes zones
            dry_zone = []
            normal_zone = []
            saturated_zone = []
            
            current_zone = None
            start_idx = 0
            
            for i, v in enumerate(values):
                if v < 70 and current_zone != "dry":
                    if current_zone is not None:
                        if current_zone == "normal":
                            normal_zone.append((start_idx, i-1))
                        elif current_zone == "saturated":
                            saturated_zone.append((start_idx, i-1))
                    current_zone = "dry"
                    start_idx = i
                elif 70 <= v <= 90 and current_zone != "normal":
                    if current_zone is not None:
                        if current_zone == "dry":
                            dry_zone.append((start_idx, i-1))
                        elif current_zone == "saturated":
                            saturated_zone.append((start_idx, i-1))
                    current_zone = "normal"
                    start_idx = i
                elif v > 90 and current_zone != "saturated":
                    if current_zone is not None:
                        if current_zone == "dry":
                            dry_zone.append((start_idx, i-1))
                        elif current_zone == "normal":
                            normal_zone.append((start_idx, i-1))
                    current_zone = "saturated"
                    start_idx = i
                    
            # Ajouter la dernière zone
            if current_zone == "dry":
                dry_zone.append((start_idx, len(values)-1))
            elif current_zone == "normal":
                normal_zone.append((start_idx, len(values)-1))
            elif current_zone == "saturated":
                saturated_zone.append((start_idx, len(values)-1))
                
            # Colorer et annoter chaque zone
            for start, end in dry_zone:
                if end > start:
                    ax.axvspan(times[start], times[end], color='orange', alpha=0.1)
                    mid_idx = (start + end) // 2
                    ax.annotate('Air sec (<70% HR)\nDessèchement possible', 
                             xy=(times[mid_idx], values[mid_idx]),
                             xytext=(0, 20),
                             textcoords='offset points',
                             ha='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="orange", alpha=0.7))
            
            for start, end in normal_zone:
                if end > start:
                    ax.axvspan(times[start], times[end], color='green', alpha=0.1)
                    mid_idx = (start + end) // 2
                    ax.annotate('Humidité normale (70-90% HR)', 
                             xy=(times[mid_idx], values[mid_idx]),
                             xytext=(0, -20),
                             textcoords='offset points',
                             ha='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="green", alpha=0.7))
            
            for start, end in saturated_zone:
                if end > start:
                    ax.axvspan(times[start], times[end], color='blue', alpha=0.1)
                    mid_idx = (start + end) // 2
                    ax.annotate('Air saturé (>90% HR)\nCondensat possible', 
                             xy=(times[mid_idx], values[mid_idx]),
                             xytext=(0, -20),
                             textcoords='offset points',
                             ha='center',
                             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="blue", alpha=0.7))
        
        # Configurer le graphique
        ax.set_xlabel("Temps (h)", fontsize=12)
        ax.set_ylabel("Humidité (%)", fontsize=12)
        ax.set_title("Évolution des paramètres de la phase liquide", fontsize=14, fontweight='bold')
        ax.grid(True, linestyle='--', alpha=0.7)
        
        # Ajouter une légende pour les courbes
        if legend_added:
            handles, labels = ax.get_legend_handles_labels()
            if handles:
                legend = ax.legend(handles, labels, fontsize=10, title="Simulations", loc="upper left")
        
        # Ajuster la mise en page
        self.liquid_figure.tight_layout(pad=2.0)
        
        # Mettre à jour le canvas - Utiliser le bon canvas pour la phase liquide
        self.liquid_canvas.draw()

if __name__ == "__main__":
    try:
        app = BioProcessApp()
        app.mainloop()
    except Exception as e:
        print(f"Erreur non gérée: {str(e)}")
    finally:
        # S'assurer que toutes les figures matplotlib sont fermées
        plt.close('all')
        # Forcer la fin du programme
        sys.exit(0)